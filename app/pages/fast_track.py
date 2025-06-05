# app/pages/fast_track.py
import os
import re
import json
from io import BytesIO
from datetime import datetime

import pandas as pd
import openpyxl
import pdfplumber
import streamlit as st
from PIL import Image
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image

from utils.sidebar_inputs import render_sidebar_inputs


# Adjust path to template relative to project structure
current_dir = os.path.dirname(os.path.abspath(__file__))
template_path = os.path.join(current_dir, "..", "template", "Template_FastTrack.xlsx")

st.title("ExPReSS for FastTrack")

with st.sidebar:
    st.page_link("main.py", label="ホーム", icon="🏠")
    st.page_link("pages/fast_track.py", label="Fast Track", icon="✈️")

# Render shared sidebar inputs
inputs = render_sidebar_inputs()
date = inputs['date']
normal_sample = inputs['normal_sample']
ep_institution = inputs['ep_institution']
ep_department = inputs['ep_department']
ep_responsible = inputs['ep_responsible']
ep_contact = inputs['ep_contact']
ep_tel = inputs['ep_tel']

uploaded_file = st.file_uploader("Upload FastTrack.pdf", type=["pdf"])

if uploaded_file is not None:
    try:
        with pdfplumber.open(uploaded_file) as pdf:
            all_text = "\n".join(
                page.extract_text() for page in pdf.pages if page.extract_text()
            )
    except Exception as e:
        st.error(f"PDFの読み込み中にエラーが発生しました: {e}")
        st.stop()

    lines = all_text.strip().split("\n")

    marker_ranges = [
        {
            "name": "Disease",
            "start_marker": "検査依頼時の情報",
            "end_marker": "疾患分類 （自動選択）",
            "columns": ["Gene_FastTrack", "Protein_FastTrack", "ID", "Gene", "Protein", "Mutation_Type", "Mutation_Allele_Frequency"]
        },
        {
            "name": "SNV_Indel",
            "start_marker": "ID 遺伝子名 アミノ酸変化 変異種別 変異アレル頻度",
            "end_marker": "構造異常",
            "columns": ["Gene_FastTrack", "Protein_FastTrack", "ID", "Gene", "Protein", "Mutation_Type", "Mutation_Allele_Frequency"]
        },
        {
            "name": "SV",
            "start_marker": "遺伝子名 サイトバンド 遺伝子名 サイトバンド",
            "end_marker": "注意事項",
            "columns": ["Gene_FastTrack", "Protein_FastTrack", "ID", "Gene_1", "Cytoband_1", "Gene_2", "Cytoband_2", "Mutation_Type"]
        }
    ]

    combined_json = {}

    for section in marker_ranges:
        start_idx, end_idx = -1, -1
        for i, line in enumerate(lines):
            if section["start_marker"] in line and start_idx == -1:
                start_idx = i + 1
            elif section["end_marker"] in line and start_idx != -1:
                end_idx = i
                break

        if start_idx != -1 and end_idx != -1:
            extracted_lines = lines[start_idx:end_idx]
            parsed_data = []

            for line in extracted_lines:
                items = re.split(r"\s+", line.strip())
                items = [item for item in items if item]

                record = {}
                for i, item in enumerate(items):
                    if i < len(section["columns"]):
                        record[section["columns"][i]] = item
                    else:
                        record[f"Extra_{i+1}"] = item

                parsed_data.append(record)

            combined_json[section["name"]] = parsed_data
        else:
            st.warning(f"{section['name']} の範囲が見つかりませんでした。")

    if combined_json:
        json_str = json.dumps(combined_json, ensure_ascii=False, indent=2)
        
        # Excel output processing
        disease_df = pd.DataFrame(combined_json.get("Disease", []))
        disease_df = disease_df.apply(lambda x: ' '.join(x.dropna().astype(str)), axis=1)
        disease = ' '.join(disease_df.tolist())
        disease_1 = disease.split('疾患名 ')[1].split(' エビデンスレベル')[0]
        disease_2 = disease.split('エビデンスレベル付与対象の ')[1]
        snv_indel_df = pd.DataFrame(combined_json.get("SNV_Indel", []))
        snv_indel_df = snv_indel_df[["Gene_FastTrack", "Protein_FastTrack", "Gene", "Protein", "Mutation_Type", "Mutation_Allele_Frequency"]]
        sv_df = pd.DataFrame(combined_json.get("SV", []))
        sv_df = sv_df[["Gene_FastTrack", "Protein_FastTrack", "Gene_1", "Cytoband_1", "Gene_2", "Cytoband_2", "Mutation_Type"]]

        if not os.path.exists(template_path):
            st.error(f"テンプレートファイルが見つかりません: {template_path}")
        else:
            try:
                workbook = openpyxl.load_workbook(template_path)
                sheet = workbook["FTReport"]

                # Insert expert panel inputs into specific cells (adjust cell references as needed)
                sheet['A37'] = '報告書作成日：' + date.strftime('%Y年%m月%d日')
                sheet['J2'] = ep_institution
                sheet['J4'] = ep_department + '/' + ep_responsible
                sheet['J38'] = ep_department + ' ' + ep_responsible
                sheet['H41'] = ep_contact + '\n' + '電話番号 ' + ep_tel + '\n' + ep_institution + ' ' + ep_department
                sheet['J7'] = disease_1
                sheet['J8'] = disease_2
                # sheet['A6'] = f"正常サンプル: {normal_sample}"
                log_path = os.path.join(current_dir, '../template/Logo.png')
                img = Image(log_path)
                img.width = 400
                img.height = 400 * img.height / img.width
                sheet.add_image(img, 'A1')

                # Insert SNV/Indel data
                if not snv_indel_df.empty:
                    for col_pos in [1, 3, 4, 6, 8, 9, 10, 12, 13, 14]:
                        if col_pos < len(snv_indel_df.columns):
                            snv_indel_df.insert(col_pos, f"Col_{col_pos}", "")

                    for r_idx, row in enumerate(snv_indel_df.itertuples(index=False), start=14):
                        for c_idx, value in enumerate(row, start=1):
                            sheet.cell(row=r_idx, column=c_idx, value=value)

                # Insert SV data
                if not sv_df.empty:
                    for col_pos in [1, 3, 4, 6, 8, 9, 11, 12, 14]:
                        if col_pos < len(sv_df.columns):
                            sv_df.insert(col_pos, f"Col_{col_pos}", "")
                    for r_idx, row in enumerate(sv_df.itertuples(index=False), start=25):
                        for c_idx, value in enumerate(row, start=1):
                            sheet.cell(row=r_idx, column=c_idx, value=value)

                # Remove empty rows
                rows_to_delete = []
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                    if all(cell.value is None for cell in row):
                        rows_to_delete.append(row[0].row)

                # 後ろから削除（インデックスのずれを防ぐ）
                for row_num in reversed(rows_to_delete):
                    sheet.delete_rows(row_num)

                # Adjust row heights and merge cells
                for row in sheet.iter_rows():
                    row_values = [cell.value for cell in row]
                    contains_expert_panel = any('Fast-track持ち回り協議結果報告書' in str(value) for value in row_values)
                    contains_summary = any('以上の遺伝子異常を確認しました。' in str(value) for value in row_values)
                    contains_postal_code = any('〒' in str(value) for value in row_values)

                    if contains_postal_code:
                        sheet.row_dimensions[row[0].row].height = 100
                        sheet.merge_cells(start_row=row[0].row, start_column=8, end_row=row[0].row, end_column=16)
                        for cell in row:
                            cell.alignment = Alignment(vertical='center', wrap_text=True)
                    elif contains_summary:
                        sheet.row_dimensions[row[0].row].height = 100
                        sheet.merge_cells(start_row=row[0].row, start_column=1, end_row=row[0].row, end_column=16)
                        for cell in row:
                            cell.alignment = Alignment(vertical='center', wrap_text=True)
                    elif contains_expert_panel:
                        sheet.row_dimensions[row[0].row].height = 100

                # 印刷範囲は'contains_postal_code'を含むセルまで
                sheet.print_area = f'A1:P{sheet.max_row}'

                output = BytesIO()
                workbook.save(output)
                workbook.close()
                output.seek(0)

                st.download_button(
                    label="Download Excel",
                    data=output,
                    file_name="FastTrack.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"Excelの出力中にエラーが発生しました: {e}")
    else:
        st.error("有効なデータが抽出されませんでした。")
        
with st.expander("## README", expanded=False):
    with open('README.md', 'r', encoding='utf-8') as f:
        readme_content = f.read()
    st.markdown(readme_content)