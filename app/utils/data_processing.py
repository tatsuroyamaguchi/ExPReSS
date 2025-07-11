import fnmatch
import glob
import json
import os
from io import BytesIO
import xml.etree.ElementTree as ET

import openpyxl
import pandas as pd
import streamlit as st

from .excel_handling import write_df_to_sheet, excel_hemesight, excel_foundationone, excel_genminetop, excel_guardant360
from .link_generator import link_generator
from .parameter import Base, Transcript, Database, Gene, Columns


def cancer_gene_census():
    cosmic_files = glob.glob(Database.COSMIC_PATH)

    if not cosmic_files:
        raise FileNotFoundError(f"No file matching pattern: {Database.COSMIC_PATH}")
    
    cancergenecensus_path = cosmic_files[0]

    df_cgc = pd.read_csv(cancergenecensus_path, sep='\t', encoding='utf-8')
    df_cgc = df_cgc[['GENE_SYMBOL', 'ROLE_IN_CANCER', 'TIER']].copy()
    df_cgc['ROLE_IN_CANCER'] = df_cgc['ROLE_IN_CANCER'].replace({'oncogene': 'OG', 'fusion': 'FU'}, regex=True).str.replace(', ', '/', regex=False)
    df_cgc['Role_Tier'] = df_cgc['ROLE_IN_CANCER'] + '[' + df_cgc['TIER'].astype(str) + ']'
    df_cgc = df_cgc[['GENE_SYMBOL', 'Role_Tier']].rename(columns={'GENE_SYMBOL': 'geneSymbol', 'Role_Tier': 'Role'})
    return df_cgc


def cancer_mutation_census():
    cosmic_files = glob.glob(Database.COSMIC_37_PATH)

    if not cosmic_files:
        raise FileNotFoundError(f"No file matching pattern: {Database.COSMIC_37_PATH}")
    
    mutationcensus_path = cosmic_files[0]

    df_cmc = pd.read_csv(mutationcensus_path, sep='\t', encoding='utf-8', compression='gzip')
    df_cmc = df_cmc[['GENE_NAME', 'Mutation CDS', 'COSMIC_SAMPLE_MUTATED']].copy()
    df_cmc = df_cmc.rename(columns={'GENE_NAME': 'geneSymbol', 'Mutation CDS': 'cdsChange', 'COSMIC_SAMPLE_MUTATED': 'COSMIC_Mutation'})
    return df_cmc


def process_hemsight(analysis_type, json_data, template_path, date, normal_sample, ep_institution, ep_department, ep_responsible, ep_contact, ep_tel):
    data = json.loads(json_data)
    wb = openpyxl.load_workbook(template_path)
    output_stream = BytesIO()

    write_df_to_sheet(data.get('testInfo', []), 'TestInfo', wb)
    write_df_to_sheet(data.get('caseData', []), 'CaseData', wb)

    short_variants = data.get('variants', {}).get('shortVariants', [])

    df_cgc = cancer_gene_census() 
    
    transcripts_data = []
    for variant in short_variants:
        for transcript in variant.get('transcripts', []):
            flat_data = {**variant, **transcript}
            flat_data.pop('transcripts', None)
            transcripts_data.append(flat_data)
 
    progress_text = st.empty()        
    for i, row in enumerate(transcripts_data):
        transcript_id = row.get('transcriptId')
        gene_symbol = row.get('geneSymbol')
        role_row = df_cgc[df_cgc['geneSymbol'] == gene_symbol]
        role_in_cancer = role_row['Role'].values[0] if not role_row.empty else ''
        row['Role_in_Cancer'] = role_in_cancer 
        amino_acids_change = row.get('aminoAcidsChange')
        cds_change = row.get('cdsChange')
        chromosome = row.get('chromosome')
        pos = row.get('position')
        ref = row.get('referenceAllele')
        alt = row.get('alternateAllele')
        gene_id = row.get('geneID')
        dbsnp = row.get('database').get('dbSNP')
        if isinstance(dbsnp, list):
            dbsnp = dbsnp[0] if dbsnp else '' 
            
        progress_text.text(f"Processing {i + 1} of {len(transcripts_data)} variants...")
        
        link_generator(analysis_type, row, gene_id, transcript_id, chromosome, pos, ref, alt, cds_change, gene_symbol, amino_acids_change, dbsnp)

    write_df_to_sheet(transcripts_data, 'ShortVariants', wb)

    rearrangements = data.get('variants', {}).get('rearrangements', [])
    breakends_data = []
    for rearrangement in rearrangements:
        for breakend in rearrangement.get('breakends', []):
            for transcript in breakend.get('transcripts', []):
                flat_data = {**rearrangement, **breakend, **transcript}
                flat_data.pop('breakends', None)
                flat_data.pop('transcripts', None)
                breakends_data.append(flat_data)
    write_df_to_sheet(breakends_data, 'Rearrangements', wb)

    write_df_to_sheet(data.get('sequencingSamples', []), 'SequencingSamples', wb)
    write_df_to_sheet(data.get('sampleIdentity', []), 'SampleIdentity', wb)
    write_df_to_sheet(data.get('tumorCellContamination', []), 'TumorContamination', wb)

    classification_match = data.get('jshGenomeGuideline', {}).get('classificationMatch', [])
    classification_data = []
    for match in classification_match:
        for level in match.get('detectedVariants', []):
            flat_data = {**match, **level}
            flat_data.pop('detectedVariants', None)
            if 'evidenceLevel' in flat_data:
                evidence_levels = flat_data.pop('evidenceLevel')
                for evidence in evidence_levels:
                    flat_data[f'{evidence["type"]}_level'] = evidence.get('level', '-')
                    flat_data[f'{evidence["type"]}_evidence'] = evidence.get('evidence', '-')
            classification_data.append(flat_data)    
    write_df_to_sheet(classification_data, 'ClassificationMatch', wb)

    fast_track = data.get('jshGenomeGuideline', {}).get('fastTrack', [])
    fast_track_data = []
    for track in fast_track:
        for level in track.get('detectedVariants', []):
            flat_data = {**track, **level}
            flat_data.pop('detectedVariants', None)
            if 'evidenceLevel' in flat_data:
                evidence_levels = flat_data.pop('evidenceLevel')
                for evidence in evidence_levels:
                    flat_data[f'{evidence["type"]}_level'] = evidence.get('level', '-')
                    flat_data[f'{evidence["type"]}_evidence'] = evidence.get('evidence', '-')
            fast_track_data.append(flat_data)
    write_df_to_sheet(fast_track_data, 'FastTrack', wb)

    wb.save(output_stream)
    output_stream.seek(0)

    output_stream = excel_hemesight(analysis_type, output_stream, date, normal_sample, ep_institution, ep_department, ep_responsible, ep_contact, ep_tel)

    def process_rearrangements(output_stream):
        df_rearrangements = pd.read_excel(output_stream, sheet_name='Rearrangements')
        df_proteinpaint = df_rearrangements.copy()
        df_proteinpaint.loc[:, 'geneSymbol'] = df_proteinpaint['geneSymbol'].replace({'D4Z4': 'DUX4'})
        transcriptID = Transcript.TRANSCRIPT_ID
        for gene, transcript in transcriptID.items():
            df_proteinpaint.loc[df_proteinpaint['geneSymbol'].str.contains(gene, case=False) & df_proteinpaint['transcriptId'].isna(), 'transcriptId'] = transcript

        df_proteinpaint.loc[:, 'geneSymbol'] = df_proteinpaint['geneSymbol'].str.split(' ').str[0].str.split('(').str[0].str.split('-').str[0]
        df_proteinpaint = df_proteinpaint.pivot_table(index='itemId', columns=df_proteinpaint.groupby('itemId').cumcount(), values=['geneSymbol', 'transcriptId', 'chromosome', 'startPosition', 'matePieceLocation'], aggfunc='first')
        df_proteinpaint.columns = [f'{col[0]}_{col[1]}' for col in df_proteinpaint.columns]
        df_proteinpaint = df_proteinpaint.reset_index()
        df_proteinpaint = df_proteinpaint.loc[:, ~df_proteinpaint.columns.str.contains('_2')]
        df_proteinpaint.columns = ['itemId', 'chr_a', 'chr_b', 'gene_a', 'gene_b', 'strand_a', 'strand_b', 'position_a', 'position_b', 'refseq_a', 'refseq_b']
        df_proteinpaint['refseq_a'] = df_proteinpaint['refseq_a'].str.split('.').str[0]
        df_proteinpaint['refseq_b'] = df_proteinpaint['refseq_b'].str.split('.').str[0]
        df_proteinpaint['strand_a'] = df_proteinpaint['strand_a'].apply(lambda x: '+' if x == 'downstream' else '-')
        df_proteinpaint['strand_b'] = df_proteinpaint['strand_b'].apply(lambda x: '+' if x == 'downstream' else '-')
        proteinpaint_path = 'proteinpaint.tsv'
        df_proteinpaint.to_csv(proteinpaint_path, sep='\t', index=False)
        
        df_disco = df_proteinpaint[Columns.PROTEINPAINT].copy()
        df_disco = df_disco[Columns.DISCO].copy()
        df_disco['chr_a'] = 'chr' + df_disco['chr_a'].astype(str)
        df_disco['chr_b'] = 'chr' + df_disco['chr_b'].astype(str)
        disco_path = 'disco.tsv'
        df_disco.to_csv(disco_path, sep='\t', index=False)
        return proteinpaint_path, disco_path
    
    proteinpaint_path, disco_path = process_rearrangements(output_stream)
    output_stream.seek(0)
    return output_stream, proteinpaint_path, disco_path


def process_foundationone(analysis_type, xml_data, template_path, date, ep_institution, ep_department, ep_responsible, ep_contact, ep_tel):
    # Parse XML data with namespace
    try:
        root = ET.fromstring(xml_data)
        ns = Base.NEME_SPACE
    except ET.ParseError as e:
        st.error(f"XML解析エラー: {e}")
        return None, None, None, None, None, None, None, None, None, None, None

    wb = openpyxl.load_workbook(template_path)
    output_stream = BytesIO()

    # Extract variant report
    variant_report = []
    for variant in root.findall('.//vr:variant-report', ns):
        variant_report.append({
            'testType': variant.get('test-type', ''),
            'gender': variant.get('gender', ''),
            'disease': variant.get('disease', ''),
            'diseaseOntology': variant.get('disease-ontology', ''),
            'tissueOfOrigin': variant.get('tissue-of-origin', ''),
            'pathologyDiagnosis': variant.get('pathology-diagnosis', ''),
            'percentTumorNuclei': variant.get('percent-tumor-nuclei', ''),
            'purityAssessment': variant.get('purity-assessment', ''),
            'specimen': variant.get('specimen', ''),
            'flowcellAnalysis': variant.get('flowcell-analysis', ''),
            'pipelineVersion': variant.get('pipeline-version', ''),
            'study': variant.get('study', ''),
            'testRequest': variant.get('test-request', ''),
        })
    write_df_to_sheet(variant_report, 'VariantReport', wb)
    
    # Extract sample
    sample_data = []
    sample_id = 1
    for sample in root.findall('.//vr:sample', ns):
        sample_data.append({
            'name': sample.get('name', ''),
            'baitSet': sample.get('bait-set', ''),
            'nucleicAcidType': sample.get('nucleic-acid-type', ''),
            'meanExonDepth': sample.get('mean-exon-depth', ''),
        })
        sample_id += 1
    write_df_to_sheet(sample_data, 'Sample', wb)
    
    # Extract QC
    qc_data = []
    for qc in root.findall('.//vr:quality-control', ns):
        qc_data.append({
            'status': qc.get('status', '')
        })
        sample_id += 1
    write_df_to_sheet(qc_data, 'QC', wb)

    df_cgc = cancer_gene_census()
    df_cmc = cancer_mutation_census()
    
    variants_data = []
    variant_id = 1
    for variant in root.findall('.//vr:short-variant', ns):
        position = variant.get('position', '')
        chromosome = position.split(':')[0].replace('chr', '') if ':' in position else ''
        pos = position.split(':')[1] if ':' in position else ''
        allele_fraction = float(variant.get('allele-fraction', '0'))
        depth = int(variant.get('depth', '0'))
        gene_symbol = variant.get('gene', '')
        gene_symbol = Gene.HUGO_SYMBOL.get(gene_symbol, gene_symbol)
        cds_change = 'c.' + variant.get('cds-effect', '')
        
        role_in_cancer = df_cgc[df_cgc['geneSymbol'] == gene_symbol]['Role'].values[0] if not df_cgc[df_cgc['geneSymbol'] == gene_symbol].empty else ''
        
        cosmic_mutation = (
            df_cmc[(df_cmc['geneSymbol'] == gene_symbol) & (df_cmc['cdsChange'] == cds_change)]['COSMIC_Mutation'].values[0]
            if not df_cmc[(df_cmc['geneSymbol'] == gene_symbol) & (df_cmc['cdsChange'] == cds_change)].empty
            else ''
        )

        # aminoacidの値を取得し、p.を追加
        amino_acid_change = variant.get('protein-effect', '')
        if amino_acid_change:
            amino_acid_change = 'p.' + amino_acid_change
            amino_acid_change = amino_acid_change.replace('p.splice site ', 'c.')
        
        # cdsの値を取得し、'splice site 'を削除してc.を追加
        cds_change = variant.get('cds-effect', '')
        if cds_change:
            cds_change = 'c.' + cds_change

        var_data = {
            'geneSymbol': gene_symbol,
            'Role_in_Cancer': role_in_cancer,
            'aminoAcidsChange': amino_acid_change,
            'cdsChange': cds_change,
            'alternateAlleleReadDepth': str(round(allele_fraction * depth)),
            'totalReadDepth': variant.get('depth', ''),
            'chromosome': chromosome,
            'position': pos,
            'transcriptId': variant.get('transcript', ''),
            'strand': variant.get('strand', ''),
            'equivocal': variant.get('equivocal', ''),
            'functional_effect': variant.get('functional-effect', ''),
            'status': variant.get('status', ''),
            'COSMIC_Mutation': cosmic_mutation,
        }
        variants_data.append(var_data)
        variant_id += 1

    progress_text = st.empty()
    for i, row in enumerate(variants_data):
        gene_symbol = row.get('geneSymbol')
        transcript_id = row.get('transcriptId')
        # gene_symbolがMUTYHの場合、transcript_idをNM_001048171.1に設定
        transcript_id_mapping = Transcript.TRANSCRIPT_ID
        if gene_symbol in transcript_id_mapping:
            transcript_id = transcript_id_mapping[gene_symbol]
    
        amino_acids_change = row.get('aminoAcidsChange')
        cds_change = row.get('cdsChange')
        chromosome = row.get('chromosome')
        pos = row.get('position')
        ref = row.get('referenceAllele')
        alt = row.get('alternateAllele')
        gene_id = row.get('geneID')
        dbsnp = row.get('dbSNP')

        progress_text.text(f"Processing {i + 1} of {len(variants_data)} variants.... φ(..)")
        
        link_generator(analysis_type, row, gene_id, transcript_id, chromosome, pos, ref, alt, cds_change, gene_symbol, amino_acids_change, dbsnp)

    write_df_to_sheet(variants_data, 'SNV_Indel', wb)
    
    # Extract CNV
    cnv_data = []
    for cnv in root.findall('.//vr:copy-number-alteration', ns):
        
        gene_symbol = cnv.get('gene', '')
        gene_symbol = Gene.HUGO_SYMBOL.get(gene_symbol, gene_symbol)
        role_in_cancer = df_cgc[df_cgc['geneSymbol'] == gene_symbol]['Role'].values[0] if not df_cgc[df_cgc['geneSymbol'] == gene_symbol].empty else ''
        
        cnv_data.append({
            'geneSymbol': gene_symbol,
            'Role_in_Cancer': role_in_cancer,
            'copyNumber': cnv.get('copy-number', ''),
            'equivocal': cnv.get('equivocal', ''),
            'numberOfExons': cnv.get('number-of-exons', ''),
            'position': cnv.get('position', ''),
            'ratio': cnv.get('ratio', ''),
            'status': cnv.get('status', ''),
            'type': cnv.get('type', '')
        })
        variant_id += 1

    write_df_to_sheet(cnv_data, 'CNV', wb)
        
    # Extract rearrangements
    rearrangements_data = []
    for rearrangement in root.findall('.//vr:rearrangement', ns):
        
        gene_symbol = rearrangement.get('targeted-gene', '')
        gene_symbol = Gene.HUGO_SYMBOL.get(gene_symbol, gene_symbol)
        role_in_cancer = df_cgc[df_cgc['geneSymbol'] == gene_symbol]['Role'].values[0] if not df_cgc[df_cgc['geneSymbol'] == gene_symbol].empty else ''
        rearrangements_data.append({
            'geneSymbol': rearrangement.get('targeted-gene', ''),
            'Role_in_Cancer': role_in_cancer,
            'alleleFraction': float(rearrangement.get('allele-fraction', '0')),
            'description': rearrangement.get('description', ''),
            'equivocal': rearrangement.get('equivocal', ''),
            'inFrame': rearrangement.get('in-frame', ''),
            'otherGene': rearrangement.get('other-gene', ''),
            'percentReads': rearrangement.get('percent-reads', ''),
            'pos1': rearrangement.get('pos1', ''),
            'pos2': rearrangement.get('pos2', ''),
            'status': rearrangement.get('status', ''),
            'supportingReadPairs': rearrangement.get('supporting-read-pairs', ''),
            'type': rearrangement.get('type', '')
        })
        variant_id += 1

    write_df_to_sheet(rearrangements_data, 'Fusion', wb)
        
    # Extract MSI
    msi_data = []
    for msi in root.findall('.//vr:microsatellite-instability', ns):
        msi_data.append({
            'status': msi.get('status', '')
        })
        variant_id += 1
    write_df_to_sheet(msi_data, 'MSI', wb)
    
    # Extract TMB
    tmb_data = []
    for tmb in root.findall('.//vr:tumor-mutation-burden', ns):
        tmb_data.append({
            'score': tmb.get('score', ''),
            'unit': tmb.get('unit', ''),
            'status': tmb.get('status', ''),
        })
        variant_id += 1
    write_df_to_sheet(tmb_data, 'TMB', wb)
        
    # Extract non-human
    non_human_data = []
    for non_human in root.findall('.//vr:non-human', ns):
        non_human_data.append({
            'organism': non_human.get('organism', ''),
            'Reads-per-million': non_human.get('reads-per-million', ''),
            'status': non_human.get('status', '')
        })
        variant_id += 1
        
    write_df_to_sheet(non_human_data, 'nonHuman', wb)
    
    wb.save(output_stream)
    output_stream.seek(0)
    
    output_stream = excel_foundationone(analysis_type, output_stream, date, ep_institution, ep_department, ep_responsible, ep_contact, ep_tel)
    output_stream.seek(0)
    return output_stream


def process_genminetop(analysis_type, xml_data, template_path, date, ep_institution, ep_department, ep_responsible, ep_contact, ep_tel):

    root = ET.fromstring(xml_data)
    wb = openpyxl.load_workbook(template_path)
    output_stream = BytesIO()

    df_cgc = cancer_gene_census()
    df_cmc = cancer_mutation_census()
    
    # Basic Information
    basic_info = []
    hospital = root.findtext('./report/owner/hospital', '')
    doctor = root.findtext('./report/owner/doctor', '')
    sex = root.findtext('./report/patient/sex', '')
    age = root.findtext('./report/patient/age', '')
    patient_id = root.findtext('./report/patient/id', '').replace('043000', '')
    ccat_id = root.findtext('./report/patient/c-cat-id', '')
    pathology = root.findtext('./report/specimen/pathology', '')
    organ = pathology.split(')_')[1].split(' -')[0]
    disease = pathology.split('_')[-1]
    germline_disclosure = root.findtext('./report/preference/germline-disclosure', '')
    germline_disclosure = '開示希望あり' if germline_disclosure == 'true' else '開示希望なし'
    
    basic_info.append({
        'hospital': hospital,
        'doctor': doctor,
        'sex': sex,
        'age': age,
        'patient_id': patient_id,
        'C_CAT_ID': ccat_id,
        'pathology': pathology,
        'organ': organ,
        'disease': disease,
        'germline_disclosure': germline_disclosure
    })
    write_df_to_sheet(basic_info, 'Sample', wb)

    reference_sections = {
        'db': ['./report/reference/db/item', ['name', 'version', 'released-at']],
        'snp-db': ['./report/reference/snp-db/item', ['name', 'version', 'genome']],
        'resource': ['./report/reference/resource/item', ['name', 'version']],
        'program': ['./report/reference/program/item', ['name', 'version']]
    }

    all_reference_data = []

    for source, (xpath, fields) in reference_sections.items():
        for item in root.findall(xpath):
            entry = {'source': source}
            for field in fields:
                entry[field.replace('-', '_')] = item.findtext(field, '')
            all_reference_data.append(entry)

    write_df_to_sheet(all_reference_data, 'DB', wb)

    # QC以下の全データを取得
    qc_data = []
    for qc in root.findall('./report/qc'):
        qc_entry = {}
        
        def extract_sample_info(qc_entry, qc, path, prefix):
            node = qc.find(path)
            if node is not None:
                qc_entry[f'{prefix}_id'] = node.findtext('id')
                sample = node.find('sample')
                if sample is not None:
                    for field in [
                        'quantity-qubit',
                        'ddcq',
                        'dv200',
                        'quantity-qubit-before-capture',
                        'quantity-qubit-after-capture',
                        'mean-length',
                        'mol-density'
                    ]:
                        key = f'{prefix}_{field.replace("-", "_")}'
                        qc_entry[key] = sample.findtext(field)
        extract_sample_info(qc_entry, qc, './library/normal/dna', 'normal')
        extract_sample_info(qc_entry, qc, './library/tumor/dna', 'tumor')
        extract_sample_info(qc_entry, qc, './library/tumor/rna', 'rna')


        # ---------- SNP Correlation ----------
        qc_entry['snp_correlation'] = qc.findtext('./sequence/snp-correlation/value')

        def extract_sequence_info(qc_entry, qc, path, prefix, depths):
            seq_node = qc.find(path)
            if seq_node is not None:
                qc_entry[f'{prefix}_seq_id'] = seq_node.findtext('id')
                qc_entry[f'{prefix}_status'] = seq_node.findtext('status')
                qc_entry[f'{prefix}_num_unique_reads'] = seq_node.findtext('num-unique-reads/value')
                qc_entry[f'{prefix}_cluster_density'] = seq_node.findtext('cluster-density/value')
                qc_entry[f'{prefix}_map_ratio'] = seq_node.findtext('map-ratio/value')
                qc_entry[f'{prefix}_num_mapped_reads'] = seq_node.findtext('num-mapped-reads/value')
                qc_entry[f'{prefix}_per_gt_q30'] = seq_node.findtext('per-gt-q30/value')
                qc_entry[f'{prefix}_num_total_reads'] = seq_node.findtext('num-total-reads/value')
                qc_entry[f'{prefix}_fragment_mean'] = seq_node.findtext('fragment/mean')
                qc_entry[f'{prefix}_fragment_sd'] = seq_node.findtext('fragment/standard-deviation')
                qc_entry[f'{prefix}_unique_read_ratio'] = seq_node.findtext('unique-read-ratio/value')
                qc_entry[f'{prefix}_cluster_pf'] = seq_node.findtext('cluster-pf/value')
                qc_entry[f'{prefix}_ontarget_ratio'] = seq_node.findtext('ontarget-ratio/value')
                qc_entry[f'{prefix}_mean_depth'] = seq_node.findtext('mean-depth/value')

                for d in depths:
                    qc_entry[f'{prefix}_cover_depth_{d}'] = seq_node.findtext(f'cover-ratio/value/depth-{d}')
                    
        extract_sequence_info(qc_entry, qc, './sequence/normal/dna', 'normal', [1, 100, 143, 200])
        extract_sequence_info(qc_entry, qc, './sequence/tumor/dna', 'tumor', [1, 100, 183, 200])


        # ---------- Tumor RNA Sequence ----------
        tumor_rna_seq = qc.find('./sequence/tumor/rna')
        if tumor_rna_seq is not None:
            qc_entry['rna_status'] = tumor_rna_seq.findtext('status')
            qc_entry['rna_seq_id'] = tumor_rna_seq.findtext('id')
            qc_entry['rna_total_reads'] = tumor_rna_seq.findtext('num-total-reads/value')
            qc_entry['rna_cluster_density'] = tumor_rna_seq.findtext('cluster-density/value')
            qc_entry['rna_cluster_pf'] = tumor_rna_seq.findtext('cluster-pf/value')
            qc_entry['rna_per_gt_q30'] = tumor_rna_seq.findtext('per-gt-q30/value')
            # ハウスキーピング遺伝子情報
            hk_gene = tumor_rna_seq.find('hk-gene')
            if hk_gene is not None:
                qc_entry['hk_num_reads'] = hk_gene.findtext('num-reads')
                qc_entry['hk_reads_per_1kbp'] = hk_gene.findtext('num-reads-per-1kbp')
                qc_entry['hk_cover_ratio'] = hk_gene.findtext('cover-ratio/value')

        # ---------- Tumor Content ----------
        qc_entry['tumor_estimated_content'] = qc.findtext('./tumor-content/estimated/value')
        qc_entry['tumor_nuclei_content'] = qc.findtext('./tumor-content/nuclei/value')

        # ---------- Append ----------
        qc_data.append(qc_entry)

    write_df_to_sheet(qc_data, 'QC', wb)
        
    # TMB
    tmb_data = []
    for tmb in root.findall('./report/result/marker/tmb'):
        for exon in tmb.findall('exon'):
            tmb_entry = {
                'num_non_synonymous_alterations': exon.findtext('num-non-synonymous-alterations', ''),
                'tmb': exon.findtext('frequency-non-synonymous-alterations', ''),
            }
            tmb_data.append(tmb_entry)
    write_df_to_sheet(tmb_data, 'TMB', wb)


    # SNV_CNV        
    variants_data_snv_indel = []
    variants_data_expression = []
    variants_germine = []
    variants_data_cnv = []
    variants_fusion = []

    # get_text 関数をループ外で定義
    def get_text(elem, tag):
        child = elem.find(tag)
        return child.text if child is not None and child.text is not None else ''

    for item in root.findall('./report/result/alterations/item'):
        variant_type = get_text(item, 'type')
        origin = get_text(item, 'origin')
        locus = get_text(item, 'locus')
        locus_split = locus.split(':') if ':' in locus else ['', '']
        chromosome = locus_split[0].replace('chr', '') if locus_split[0] else ''
        position = locus_split[1] if len(locus_split) > 1 else ''
        gene_symbol = get_text(item, 'gene')
        gene_symbol = Gene.HUGO_SYMBOL.get(gene_symbol, gene_symbol)
        transcript_id = get_text(item, 'transcript')
        # gene_symbolがMUTYHの場合、transcript_idをNM_001048171.1に設定
        if gene_symbol == 'MUTYH':
            transcript_id = 'NM_001048171.1'
        role_in_cancer = (
            df_cgc[df_cgc['geneSymbol'] == gene_symbol]['Role'].values[0]
            if not df_cgc[df_cgc['geneSymbol'] == gene_symbol].empty else ''
        )

        # breakpointの下のディレクトリにitem要素が２つある場合、itemの下にあるregion, index, length, gene, transcript, chr, posをそれぞれ_1, _2として取得
        region_1     = item.findtext("breakpoint/item[1]/region", default="")
        region_2     = item.findtext("breakpoint/item[2]/region", default="")
        index_1      = item.findtext("breakpoint/item[1]/index", default="")
        index_2      = item.findtext("breakpoint/item[2]/index", default="")
        length_1     = item.findtext("breakpoint/item[1]/length", default="")
        length_2     = item.findtext("breakpoint/item[2]/length", default="")
        gene_1       = item.findtext("breakpoint/item[1]/gene", default="")
        gene_2       = item.findtext("breakpoint/item[2]/gene", default="")
        transcript_1 = item.findtext("breakpoint/item[1]/transcript", default="")
        transcript_2 = item.findtext("breakpoint/item[2]/transcript", default="")
        chr_1        = item.findtext("breakpoint/item[1]/chr", default="")
        chr_2        = item.findtext("breakpoint/item[2]/chr", default="")
        pos_1        = item.findtext("breakpoint/item[1]/pos", default="")
        pos_2        = item.findtext("breakpoint/item[2]/pos", default="")

        variant = {
            'geneSymbol': gene_symbol,
            'transcriptId': transcript_id,
            'Role_in_Cancer': role_in_cancer,
            'chromosome': chromosome,
            'position': position,
            'referenceAllele': get_text(item, 'ref'),
            'alternateAllele': get_text(item, 'alt'),
            'transcriptId': get_text(item, 'transcript'),
            'aminoAcidsChange': get_text(item, 'protein-alteration'),
            'cdsChange': get_text(item, 'coding-dna-alteration'),
            'origin': origin,
            'alternateAlleleFrequency': get_text(item, 'allele-frequency'),
            'status': get_text(item, 'status'),
            'type': variant_type,
            'cytoband': get_text(item, 'cytoband'),
            'copyNumber': get_text(item, 'num-copy'),
            'ratio': get_text(item, 'ratio'),
            'frame': get_text(item, 'frame'),
            'gene_1': gene_1,
            'gene_2': gene_2,
            'transcript_1': transcript_1,
            'transcript_2': transcript_2,
            'region_1': region_1,
            'region_2': region_2,
            'index_1': index_1,
            'index_2': index_2,
            'length_1': length_1,
            'length_2': length_2,
            'chr_1': chr_1,
            'chr_2': chr_2,
            'pos_1': pos_1,
            'pos_2': pos_2,
        }
        if origin.lower() == 'germline':
            variants_germine.append(variant)
        elif variant_type.lower() == 'expression':
            variants_data_expression.append(variant)
        elif variant_type.lower() == 'fusion':
            variants_fusion.append(variant)
        elif 'cnv' in variant_type.lower():
            variants_data_cnv.append(variant)
        else:
            variants_data_snv_indel.append(variant)

    # SNV/Indel の処理
    def process_variants(variants, analysis_type, link_generator, label=""):
        progress_text = st.empty()
        for i, row in enumerate(variants):
            progress_text.text(f"Processing {i + 1} of {len(variants)} variants.... φ(..) {label}")

            link_generator(
                analysis_type,
                row,
                row.get('geneID', ''),
                row.get('transcriptId', ''),
                row.get('chromosome', ''),
                row.get('position', ''),
                row.get('referenceAllele', ''),
                row.get('alternateAllele', ''),
                row.get('cdsChange', ''),
                row.get('geneSymbol', ''),
                row.get('aminoAcidsChange', ''),
                row.get('dbSNP', '')
            )
    process_variants(variants_data_snv_indel, analysis_type, link_generator, label="SNV/INDEL")
    process_variants(variants_germine, analysis_type, link_generator, label="Germline")
    
    # データをExcelシートへ出力
    if variants_data_snv_indel:
        write_df_to_sheet(variants_data_snv_indel, 'SNV_Indel', wb)
    if variants_data_cnv:
        write_df_to_sheet(variants_data_cnv, 'CNV', wb)
    if variants_fusion:
        write_df_to_sheet(variants_fusion, 'Fusion', wb)
    if variants_data_expression:
        write_df_to_sheet(variants_data_expression, 'EXP', wb)
    if variants_germine:
        write_df_to_sheet(variants_germine, 'Germline', wb)
        
    # ファイル保存
    wb.save(output_stream)
    output_stream.seek(0)

    output_stream = excel_genminetop(analysis_type, output_stream, date, ep_institution, ep_department, ep_responsible, ep_contact, ep_tel)
    output_stream.seek(0)
    return output_stream


def process_guardant360(analysis_type, xlsx_data, template_path, date, ep_institution, ep_department, ep_responsible, ep_contact, ep_tel):
    wb = openpyxl.load_workbook(template_path)

    df_cgc = cancer_gene_census()
    df_cmc = cancer_mutation_census()

    df_snv = pd.read_excel(xlsx_data, sheet_name='SNV')
    df_snv = df_snv[df_snv['call'] == 1]
    
    if not df_snv.empty:
        df_snv[['referenceAllele', 'alternateAllele']] = df_snv['mut_nt'].str.split('>', expand=True)        
        df_snv.columns = ['geneSymbol', 'chromosome', 'position', 'mut_nt', 'aminoAcidsChange', 'cdsChange', 'alternateAlleleFrequency', 'call', 'transcriptId', 'exon', 'reporting_category', 'rm_reportable', 'referenceAllele', 'alternateAllele']
        df_snv.reset_index(drop=True, inplace=True)
        df_snv['aminoAcidsChange'] = df_snv['aminoAcidsChange'].apply(lambda x: "p." + str(x) if pd.notnull(x) else '')
        # 'geneID', 'dbSNP'は元のデータに存在しないため、空の列を追加
        df_snv['geneID'] = ''
        df_snv['dbSNP'] = ''
        df_snv['Role_in_Cancer'] = ''
        df_snv['status'] = ''
        df_snv['geneSymbol'] = df_snv['geneSymbol'].map(Gene.HUGO_SYMBOL).fillna(df_snv['geneSymbol'])
        
        df_snv.loc[(df_snv['rm_reportable'] == 0), 'status'] = 'LV4'
        df_snv.loc[(df_snv['rm_reportable'] == 1) & (df_snv['geneSymbol'] == 'KRAS') & (df_snv['aminoAcidsChange'] == 'G12C'), 'status'] = 'LV1'
        df_snv.loc[(df_snv['rm_reportable'] == 1) & ((df_snv['geneSymbol'] != 'KRAS') | (df_snv['aminoAcidsChange'] != 'G12C')), 'status'] = 'LV2'
        
        progress_text = st.empty()
        for i, row in df_snv.iterrows():
            gene_symbol = row['geneSymbol']
            gene_symbol = Gene.HUGO_SYMBOL.get(gene_symbol, gene_symbol)
            role_in_cancer = df_cgc[df_cgc['geneSymbol'] == gene_symbol]['Role'].values[0] if not df_cgc[df_cgc['geneSymbol'] == gene_symbol].empty else ''
            df_snv.at[i, 'Role_in_Cancer'] = role_in_cancer
            progress_text.text(f"Processing {i + 1} of {len(df_snv)} variants.... φ(..)")
            results = link_generator(analysis_type, row, row['geneID'], row['transcriptId'], row['chromosome'], row['position'], row['referenceAllele'], row['alternateAllele'], row['cdsChange'], gene_symbol, row['aminoAcidsChange'], row['dbSNP'])
            for key, value in results.items():
                df_snv.at[i, key] = value
        write_df_to_sheet(df_snv, 'SNV', wb)

    df_indel = pd.read_excel(xlsx_data, sheet_name='Indels')
    df_indel = df_indel[df_indel['call'] == 1]
    
    if not df_indel.empty:
        df_indel[['referenceAllele', 'alternateAllele']] = df_snv['mut_nt'].str.split('>', expand=True)  
        df_indel.columns = ['geneSymbol', 'chromosome', 'position', 'mut_nt', 'aminoAcidsChange', 'cdsChange', 'length', 'exon', 'type', 'alternateAlleleFrequency', 'call', 'transcriptId', 'reporting_category', 'mut_aa_short', 'rm_reportable', 'referenceAllele', 'alternateAllele']
        df_indel.reset_index(drop=True, inplace=True)
        df_indel['aminoAcidsChange'] = df_indel['aminoAcidsChange'].apply(lambda x: "p." + str(x) if pd.notnull(x) else '')
        # 'geneID', 'dbSNP'は元のデータに存在しないため、空の列を追加
        df_indel['geneID'] = ''
        df_indel['dbSNP'] = ''
        df_indel['Role_in_Cancer'] = ''
        df_indel['status'] = ''
        df_indel['geneSymbol'] = df_indel['geneSymbol'].map(Gene.HUGO_SYMBOL).fillna(df_indel['geneSymbol'])
        
        df_indel.loc[(df_indel['rm_reportable'] == 0), 'status'] = 'LV4'
        df_indel.loc[(df_indel['rm_reportable'] == 1) & (df_indel['geneSymbol'] == 'KRAS') & (df_indel['aminoAcidsChange'] == 'G12C'), 'status'] = 'LV1'
        df_indel.loc[(df_indel['rm_reportable'] == 1) & ((df_indel['geneSymbol'] != 'KRAS') | (df_indel['aminoAcidsChange'] != 'G12C')), 'status'] = 'LV2'
        progress_text = st.empty()
        for i, row in df_indel.iterrows():
            gene_symbol = row['geneSymbol']
            gene_symbol = Gene.HUGO_SYMBOL.get(gene_symbol, gene_symbol)
            role_in_cancer = df_cgc[df_cgc['geneSymbol'] == gene_symbol]['Role'].values[0] if not df_cgc[df_cgc['geneSymbol'] == gene_symbol].empty else ''
            df_indel.at[i, 'Role_in_Cancer'] = role_in_cancer
            progress_text.text(f"Processing {i + 1} of {len(df_indel)} variants.... φ(..)")
            results = link_generator(analysis_type, row, row['geneID'], row['transcriptId'], row['chromosome'], row['position'], row['referenceAllele'], row['alternateAllele'], row['cdsChange'], gene_symbol, row['aminoAcidsChange'], row['dbSNP'])
            for key, value in results.items():
                df_indel.at[i, key] = value
        write_df_to_sheet(df_indel, 'Indels', wb)
    
    df_cnv = pd.read_excel(xlsx_data, sheet_name='CNAs')
    df_cnv.columns = ['chromosome', 'geneSymbol', 'copyNumber', 'call']
    df_cnv = df_cnv[df_cnv['call'] != 0]
    
    if not df_cnv.empty:
        df_cnv['geneSymbol'] = df_cnv['geneSymbol'].map(Gene.HUGO_SYMBOL).fillna(df_cnv['geneSymbol'])
        df_cnv['Role_in_Cancer'] = ''
        df_cnv['status'] = ''
        df_cnv['type'] = ''
        
        progress_text = st.empty()
        for i, row in df_cnv.iterrows():
            gene_symbol = row['geneSymbol']
            gene_symbol = Gene.HUGO_SYMBOL.get(gene_symbol, gene_symbol)
            role_in_cancer = df_cgc[df_cgc['geneSymbol'] == gene_symbol]['Role'].values[0] if not df_cgc[df_cgc['geneSymbol'] == gene_symbol].empty else ''
            df_cnv.at[i, 'Role_in_Cancer'] = role_in_cancer

        df_cnv['type'] = df_cnv['call'].apply(lambda x: 'Amplification' if x != 0 else 'Equivocal')
        
    if not df_cnv.empty:
        df_cnv.loc[(df_cnv['call'] == 1) & (df_cnv['geneSymbol'].isin(['ERBB2', 'MET'])), 'status'] = 'LV1'
        df_cnv.loc[(df_cnv['call'] == 1) & (~df_cnv['geneSymbol'].isin(['ERBB2', 'MET'])), 'status'] = 'LV3'
        df_cnv.loc[(df_cnv['call'] == 2) & (df_cnv['geneSymbol'].isin(['ERBB2', 'MET'])), 'status'] = 'LV2'
        df_cnv.loc[(df_cnv['call'] == 2) & (~df_cnv['geneSymbol'].isin(['ERBB2', 'MET'])), 'status'] = 'LV3'
        df_cnv.loc[df_cnv['call'] == 3, 'status'] = 'Aneuploidy'
    write_df_to_sheet(df_cnv, 'CNAs', wb)
    
    df_fusion = pd.read_excel(xlsx_data, sheet_name='Fusions')
    df_fusion = df_fusion[df_fusion['call'] == 1]
    
    if not df_fusion.empty:
        df_fusion['gene_a'] = df_fusion['gene_a'].map(Gene.HUGO_SYMBOL).fillna(df_fusion['gene_a'])
        df_fusion['gene_b'] = df_fusion['gene_b'].map(Gene.HUGO_SYMBOL).fillna(df_fusion['gene_b'])
        df_fusion['Role_in_Cancer'] = ''
        df_fusion['status'] = ''
        df_fusion['type'] = 'Fusion'
        
        progress_text = st.empty()
        for i, row in df_fusion.iterrows():
            gene_symbol = row['gene_a']
            gene_symbol = Gene.HUGO_SYMBOL.get(gene_symbol, gene_symbol)
            role_in_cancer = df_cgc[df_cgc['geneSymbol'] == gene_symbol]['Role'].values[0] if not df_cgc[df_cgc['geneSymbol'] == gene_symbol].empty else ''
            df_fusion.at[i, 'Role_in_Cancer'] = role_in_cancer
            
        df_fusion['geneSymbol'] = df_fusion['gene_a'] + ' - ' + df_fusion['gene_b']
        df_fusion['type'] = 'Fusion'
    
    if not df_fusion.empty:
        df_fusion.loc[df_fusion['gene_a'].isin(['ALK', 'NTRK1', 'RET', 'ROS1']), 'status'] = 'LV2'
        df_fusion.loc[df_fusion['gene_a'].isin(['FGFR2', 'FGFR3']), 'status'] = 'LV3'
    write_df_to_sheet(df_fusion, 'Fusions', wb)
    
    df_msi = pd.read_excel(xlsx_data, sheet_name='MSI')
    write_df_to_sheet(df_msi, 'MSI', wb)
    
    df_qc = pd.read_excel(xlsx_data, sheet_name='QC')
    write_df_to_sheet(df_qc, 'QC', wb)
    
    output_stream = BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    output_stream = excel_guardant360(analysis_type, output_stream, date, ep_institution, ep_department, ep_responsible, ep_contact, ep_tel)
    output_stream.seek(0)
    return output_stream
