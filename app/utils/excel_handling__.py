import io
import os
import re
import json
from copy import copy
from io import BytesIO

import pandas as pd
import pdfplumber
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
import streamlit as st

from .parameter import Abbreviation, Database, Columns, ReadRange


def write_df_to_sheet(data_section, sheet_name, wb):
    # データを正規化（既にDataFrameならそのまま）
    if not isinstance(data_section, pd.DataFrame):
        df = pd.json_normalize(data_section)
    else:
        df = data_section
    # シートが存在するか確認し、なければ新規作成
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.create_sheet(title=sheet_name)
    # DataFrame を Excel シートに書き込み
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            if isinstance(value, list):
                value = "\n".join(map(str, value))
            elif isinstance(value, dict):
                value = json.dumps(value, ensure_ascii=False)
            sheet.cell(row=r_idx, column=c_idx, value=value)

def add_logo(current_dir, sheet, cell):
    logo_path = os.path.join(current_dir, Database.LOGO_PATH)
    if os.path.exists(logo_path):
        img = Image(logo_path)
        orig_w, orig_h = img.width, img.height
        img.width = 400
        img.height = int(400 * orig_h / orig_w)
        sheet.add_image(img, cell)

def delete_blank_lines(sheet):
    for row in range(sheet.max_row, 0, -1):
        if all(cell.value is None for cell in sheet[row]):
            sheet.delete_rows(row, 1)

def insert_row(wb, df, sheet_name, start_row, start_col, end_col):
    if df is None or df.empty:
        return
    sheet = wb[sheet_name]
    # スタイルを取得
    select_range = f'{start_col}{start_row}:{end_col}{start_row}'
    styles = []
    for cell in sheet[select_range][0]:
        style = {
            'column': copy(cell.column),
            'font': copy(cell.font),
            'border': copy(cell.border),
            'fill': copy(cell.fill),
            'number_format': copy(cell.number_format),
            'alignment': copy(cell.alignment)
        }
        styles.append(style)
    # 行を挿入
    insert_count = len(df) - 1
    if insert_count > 0:
        sheet.insert_rows(start_row, amount=insert_count)
    # スタイルを適用
    for r_idx in range(start_row, start_row + insert_count):
        for c_idx, style in enumerate(styles, start=1):
            cell = sheet.cell(row=r_idx, column=c_idx)
            cell.font = style['font']
            cell.border = style['border']
            cell.fill = style['fill']
            cell.number_format = style['number_format']
            cell.alignment = style['alignment']

def excel_hemesight(analysis_type, output_stream, date, normal_sample, ep_institution, ep_department, ep_responsible, ep_contact, ep_tel):
    current_dir = os.getcwd()
    
    df_short = pd.read_excel(output_stream, sheet_name='ShortVariants')
    df_short = df_short.copy()
    df_short['aminoAcidsChange'] = df_short['aminoAcidsChange'].fillna(df_short['cdsChange'])
    df_short = df_short.drop(columns=['cdsChange'])
    df_short['supportingReadCount'] = df_short['alternateAlleleReadDepth'].replace(0, pd.NA)
    df_short['totalReadDepth'] = df_short['totalReadDepth'].replace(0, pd.NA)
    df_short['alternateAlleleFrequency'] = df_short.apply(
        lambda x: x['alternateAlleleReadDepth'] / x['totalReadDepth'] 
        if pd.notna(x['alternateAlleleReadDepth']) and pd.notna(x['totalReadDepth']) else pd.NA, axis=1)
    df_short['alternateAlleleFrequency'] = df_short['alternateAlleleFrequency'].apply(
        lambda x: f"{x*100:.1f}%" if pd.notna(x) else '-')

    df_short = df_short[Columns.HEMESIGHT_SHORT]

    df_short['database.gnomAD'] = df_short['database.gnomAD'].apply(lambda x: f"*{x}" if x != '-' else x)
    df_short.loc[df_short['database.tommo'] == '-', 'database.tommo'] = df_short['database.gnomAD']
    df_short = df_short.drop(columns=['database.gnomAD'])
    df_short.insert(3, 'a', ''); df_short.insert(4, 'b', ''); df_short.insert(8, 'd', ''); df_short.insert(10, 'e', '')

    df_gl = df_short[df_short['itemId'].str.startswith('GL')].drop(columns=['itemId'])
    df_mt = df_short[df_short['itemId'].str.startswith('MT')].drop(columns=['itemId'])
    df_mnv = df_short[df_short['itemId'].str.startswith('MNV')].drop(columns=['itemId'])
    df_up = df_short[df_short['itemId'].str.startswith('UP')].drop(columns=['itemId'])

    df_fast = df_short[['itemId', 'geneSymbol', 'aminoAcidsChange']]
    df_fast = df_fast[
        ((df_fast['geneSymbol'] == 'FLT') & (df_fast['aminoAcidsChange'].str.contains('Asp835'))) |
        ((df_fast['geneSymbol'] == 'BRAF') & (df_fast['aminoAcidsChange'].str.contains('Val600Glu'))) |
        ((df_fast['geneSymbol'] == 'EZH2') & (df_fast['aminoAcidsChange'].str.contains('Tyr641'))) |
        ((df_fast['geneSymbol'] == 'IDH1') & (df_fast['aminoAcidsChange'].str.contains('Arg132Cys|Arg132Leu|Arg132His')))
    ]

    df_fast['itemId'] = '変異あり'
    
    required_columns = ['gene', 'cancerType', 'chromosomalChange', 'fastTrackVariant', 'comments', 'analysisType']
    try:
        df_fast_track = pd.read_excel(output_stream, sheet_name='FastTrack')
        missing = [col for col in required_columns if col not in df_fast_track.columns]
        for col in missing:
            df_fast_track[col] = pd.NA
        df_fast_track = df_fast_track[required_columns]
    except Exception:
        df_fast_track = pd.DataFrame(columns=required_columns)

    ####################################
    wb = openpyxl.load_workbook(output_stream)
    sheet = wb['FTReport']    

    gene_row_map = {'FLT': 12, 'BRAF': 13, 'EZH2': 14, 'IDH1': 15}

    for gene, group in df_fast.groupby('geneSymbol'):
        if gene in gene_row_map:
            row_num = gene_row_map[gene]
            row_data = group.iloc[0].tolist()
            for c_idx, value in enumerate(row_data, 1):
                sheet.cell(row=row_num, column=c_idx, value=value)
 
    for row in sheet.iter_rows():
        row_values = [cell.value for cell in row]
        contains_postal_code = any('〒' in str(value) for value in row_values)

        if contains_postal_code:
            sheet.row_dimensions[row[0].row].height = 100
            sheet.merge_cells(start_row=row[0].row, start_column=10, end_row=row[0].row, end_column=16)
            for cell in row:
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                
    # 基本情報の設定
    add_logo(current_dir, sheet, cell='A1')
    sheet['D2'] = analysis_type
    sheet['A22'] = '報告書作成日：' + date.strftime('%Y年%m月%d日')
    sheet['J2'] = ep_institution
    sheet['A23'] = ep_institution + ' ' + ep_department
    sheet['J4'] = ep_department + ' / ' + ep_responsible
    sheet['J23'] = ep_responsible
    sheet['J26'] = ep_institution + ' ' + ep_department + '\n' + ep_contact + '\n' + '電話番号 ' + ep_tel

    ####################################
    # 1. データ読み込み
    df_rearrangements = pd.read_excel(output_stream, sheet_name='Rearrangements')
    df_rearrange = df_rearrangements[Columns.HEMESIGHT_REARRANGEMENT].copy()

    # 2. geneSymbolの前処理
    df_rearrange['geneSymbol'] = df_rearrange['geneSymbol'].fillna('-')
    df_rearrange['geneSymbol'] = df_rearrange.apply(
        lambda row: f"{row['geneSymbol']} [ex {int(row['number.number'])}]" 
        if pd.notna(row['number.number']) else row['geneSymbol'], axis=1)
    df_rearrange = df_rearrange.drop(columns=['number.number'])
    
    # matePieceLocationがupstreamの場合(+)、downstreamの場合(-)に変換
    df_rearrange['matePieceLocation'] = df_rearrange['matePieceLocation'].replace({'upstream': '(-)', 'downstream': '(+)'})

    # 3. 共通処理用の関数定義
    def process_rearrangement_data(df, prefix, columns, check_value, include_inserted_sequence=True, include_read_count=True):
        if df.empty:
            result = pd.DataFrame(columns=columns)
            result['Check'] = check_value
            return result

        # geneSymbolに位置情報を追加
        df.loc[:, 'geneSymbol'] = (
            df['geneSymbol'] + '\n' +
            df['chromosome'].astype(str) + ':' +
            df['startPosition'].astype(str) + ' ' +
            df['matePieceLocation'].astype(str)
        )

        # ピボットテーブル作成
        df_pivot = df.pivot_table(
            index=['itemId', 'chromosome', 'startPosition'],
            columns=df.groupby(['itemId', 'chromosome', 'startPosition']).cumcount(),
            values='geneSymbol',
            aggfunc='first'
        )
        df_pivot.columns = [f'geneSymbol_{col}' for col in df_pivot.columns]
        df_pivot = df_pivot.reset_index()

        # geneSymbolの結合
        df_pivot['geneSymbol'] = df_pivot.apply(
            lambda row: f"{row['geneSymbol_0']}/{row['geneSymbol_1']}" 
            if 'geneSymbol_1' in df_pivot.columns and pd.notna(row['geneSymbol_1']) 
            else row['geneSymbol_0'], axis=1
        )
        df_pivot = df_pivot.drop(columns=[col for col in ['geneSymbol_0', 'geneSymbol_1'] if col in df_pivot.columns])

        # マージとデータ整形
        merge_cols = ['itemId', 'chromosome', 'startPosition', 'rearrangementType', 'function.mitelman']
        if include_inserted_sequence:
            merge_cols.append('insertedSequence')
        if include_read_count:
            merge_cols.append('supportingReadCount')
        df_merged = pd.merge(df_pivot, df_rearrange[merge_cols], on=['itemId', 'chromosome', 'startPosition'])
        df_merged['function.mitelman'] = df_merged['function.mitelman'].fillna('-')
        df_merged['supportingReadCount'] = df_merged['supportingReadCount'].astype(str)

        # ピボットテーブル再作成
        pivot_values = ['geneSymbol', 'rearrangementType', 'function.mitelman']
        if include_inserted_sequence:
            pivot_values.append('insertedSequence')
        if include_read_count:
            pivot_values.append('supportingReadCount')
        df_final = df_merged.drop_duplicates().pivot_table(
            index='itemId',
            columns=df_merged.groupby('itemId').cumcount(),
            values=pivot_values,
            aggfunc='first'
        )
        df_final.columns = [f'{col[0]}_{col[1]}' for col in df_final.columns]
        df_final = df_final.reset_index().reindex(columns=columns, fill_value='')
        df_final['Check'] = check_value
        return df_final

    # 4. SV, UP_SV, FU, DUの処理
    df_sv = process_rearrangement_data(
        df_rearrange[df_rearrange['itemId'].str.startswith('SV')],
        'SV', Columns.HEMESIGHT_SV, 'SV'
    )
    df_up_sv = process_rearrangement_data(
        df_rearrange[df_rearrange['itemId'].str.startswith('UP')],
        'UP', Columns.HEMESIGHT_SV, 'UP_SV'
    )
    df_fu = process_rearrangement_data(
        df_rearrange[df_rearrange['itemId'].str.startswith('FU')],
        'FU', Columns.HEMESIGHT_FU, 'FU_RNA', include_inserted_sequence=False, include_read_count=True
    )
    df_du = process_rearrangement_data(
        df_rearrange[df_rearrange['itemId'].str.startswith('DU')],
        'DU', Columns.HEMESIGHT_SV, 'DU'
    )

    # 5. CaseDataとJSHデータの処理
    df_case_data = pd.read_excel(output_stream, sheet_name='CaseData')
    cancer_type = next(
        (abbr for key, abbr in Abbreviation.ABBR_DISEASE_NAME.items() if key in df_case_data['cancerType'].values[0]),
        'Other'
    )

    # 6. 全遺伝子の収集
    all_genes = set(df_gl['geneSymbol']).union(
        df_mt['geneSymbol'], df_mnv['geneSymbol'], df_up['geneSymbol'],
        df_sv[['geneSymbol_0', 'geneSymbol_1']].stack(),
        df_up_sv[['geneSymbol_0', 'geneSymbol_1']].stack(),
        df_fu[['geneSymbol_0', 'geneSymbol_1']].stack(),
        df_du[['geneSymbol_0', 'geneSymbol_1']].stack()
    )
    all_genes = {gene.split('\n')[0].split(' [')[0] for gene in all_genes}

    # 7. JSHデータのフィルタリング
    jsh_path = os.path.join(current_dir, Database.JSA_PATH)
    df_jsh = pd.read_csv(jsh_path, encoding='utf-8')
    df_jsh = df_jsh[df_jsh['Disease'].str.contains(cancer_type) & df_jsh['Gene'].isin(all_genes)].sort_values('Gene')

    # 8. JSHエビデンスと薬剤データの作成
    df_jsh_evidence = df_jsh[Columns.HEMESIGHT_JSA].copy()
    df_jsh_evidence['Check'] = 'evidence'
    df_jsh_drugs = df_jsh[['Gene', 'Drugs', 'Comments']].copy()
    df_jsh_drugs['Check'] = 'drug'
    
    ####################################    
    sheet = wb['Report']

    # 基本情報の設定
    add_logo(current_dir, sheet, cell='A1')
    sheet['D2'] = analysis_type
    sheet['D4'] = date.strftime('%Y年%m月%d日')
    sheet['A33'] = '報告書作成日：' + date.strftime('%Y年%m月%d日')
    sheet['F11'] = normal_sample
    sheet['J2'] = ep_institution
    sheet['A34'] = ep_institution + ' ' + ep_department
    sheet['J4'] = ep_department + ' / ' + ep_responsible
    sheet['J34'] = ep_responsible
    sheet['J37'] = ep_institution + ' ' + ep_department + '\n' + ep_contact + '\n' + '電話番号 ' + ep_tel

    start_row_jsh_evidence = 15
    insert_row(wb, df_jsh_evidence, sheet_name='Report', start_row=start_row_jsh_evidence, start_col='A', end_col='P')
    for insert_pos in [2, 3, 5, 7, 8, 9, 11, 12, 13, 15]:
        df_jsh_evidence.insert(insert_pos, chr(96 + insert_pos), '')
    start_row_mt = start_row_jsh_evidence + 31 + len(df_jsh_drugs) - 1 if len(df_jsh_drugs) > 1 else start_row_jsh_evidence + 31
    insert_row(wb, df_mt, sheet_name='Report', start_row=start_row_mt, start_col='A', end_col='P')
    start_row_mnv = start_row_mt + 4 +len(df_mt) - 1 if len(df_mt) > 1 else start_row_mt + 4
    insert_row(wb, df_mnv, sheet_name='Report', start_row=start_row_mnv, start_col='A', end_col='P')
    start_row_up = start_row_mnv + 4 + len(df_mnv) - 1 if len(df_mnv) > 1 else start_row_mnv + 4
    insert_row(wb, df_up, sheet_name='Report', start_row=start_row_up, start_col='A', end_col='P')
    start_row_sv = start_row_up + 4 + len(df_up) - 1 if len(df_up) > 1 else start_row_up + 4
    insert_row(wb, df_sv, sheet_name='Report', start_row=start_row_sv, start_col='A', end_col='P')
    for insert_pos in [1, 2, 4, 5, 7, 8, 10, 12, 13, 14, 15]:
        df_sv.insert(insert_pos, chr(96 + insert_pos), '')
    start_row_up_sv = start_row_sv + 4 + len(df_sv) - 1 if len(df_sv) > 1 else start_row_sv + 4
    insert_row(wb, df_up_sv, sheet_name='Report', start_row=start_row_up_sv, start_col='A', end_col='P')
    for insert_pos in [1, 2, 4, 5, 7, 8, 10, 12, 13, 14, 15]:
        df_up_sv.insert(insert_pos, chr(96 + insert_pos), '')
    start_row_fu = start_row_up_sv + 4 + len(df_up_sv) - 1 if len(df_up_sv) > 1 else start_row_up_sv + 4
    insert_row(wb, df_fu, sheet_name='Report', start_row=start_row_fu, start_col='A', end_col='P')
    for insert_pos in [1, 2, 4, 5, 7, 8, 10, 12, 13, 14, 15]:
        df_fu.insert(insert_pos, chr(96 + insert_pos), '')
    start_row_du = start_row_fu + 4 + len(df_fu) - 1 if len(df_fu) > 1 else start_row_fu + 4
    insert_row(wb, df_du, sheet_name='Report', start_row=start_row_du, start_col='A', end_col='P')
    for insert_pos in [1, 2, 4, 5, 7, 8, 10, 12, 13, 14, 15]:
        df_du.insert(insert_pos, chr(96 + insert_pos), '')
    start_row_gl = start_row_du + 4 + len(df_du) - 1 if len(df_du) > 1 else start_row_du + 4
    insert_row(wb, df_gl, sheet_name='Report', start_row=start_row_gl, start_col='A', end_col='P')

    for df_section, start_row in [(df_jsh_evidence, start_row_jsh_evidence),
                                  (df_mt, start_row_mt), (df_mnv, start_row_mnv), (df_up, start_row_up), 
                                  (df_sv, start_row_sv), (df_up_sv, start_row_up_sv), (df_fu, start_row_fu), 
                                  (df_du, start_row_du), (df_gl, start_row_gl)
                                  ]:
        for r_idx, row in enumerate(dataframe_to_rows(df_section, index=False, header=False), start_row):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)
                sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(wrap_text=False, vertical='top')

    for row in sheet.iter_rows():
        row_values = [cell.value for cell in row]
        contains_expert_panel = any('エキスパートパネルレポート' in str(value) for value in row_values)
        contains_tohoku = any('Tohoku' in str(value) for value in row_values)
        contains_postal_code = any('〒' in str(value) for value in row_values)
        check_sv = 'SV' in str(row[16].value)
        check_up_sv = 'UP_SV' in str(row[16].value)
        check_fu = 'FU_RNA' in str(row[16].value)
        check_du = 'DU' in str(row[16].value)
        check_evidence = 'evidence' in str(row[16].value)
        check_drug = 'drug' in str(row[16].value)

        if check_sv or check_up_sv or check_du:
            sheet.row_dimensions[row[0].row].height = 50
            sheet.merge_cells(start_row=row[0].row, start_column=1, end_row=row[0].row, end_column=3)
            sheet.merge_cells(start_row=row[0].row, start_column=4, end_row=row[0].row, end_column=6)
            sheet.merge_cells(start_row=row[0].row, start_column=7, end_row=row[0].row, end_column=9)
            sheet.merge_cells(start_row=row[0].row, start_column=10, end_row=row[0].row, end_column=11)
            sheet.merge_cells(start_row=row[0].row, start_column=12, end_row=row[0].row, end_column=14)
            for cell in row:
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        elif check_fu:
            sheet.row_dimensions[row[0].row].height = 50
            sheet.merge_cells(start_row=row[0].row, start_column=1, end_row=row[0].row, end_column=3)
            sheet.merge_cells(start_row=row[0].row, start_column=4, end_row=row[0].row, end_column=6)
            sheet.merge_cells(start_row=row[0].row, start_column=7, end_row=row[0].row, end_column=9)
            sheet.merge_cells(start_row=row[0].row, start_column=10, end_row=row[0].row, end_column=11)
            sheet.merge_cells(start_row=row[0].row, start_column=12, end_row=row[0].row, end_column=14)
            for cell in row:
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        elif check_evidence:
            sheet.row_dimensions[row[0].row].height = 125
            sheet.merge_cells(start_row=row[0].row, start_column=2, end_row=row[0].row, end_column=4)
            sheet.merge_cells(start_row=row[0].row, start_column=5, end_row=row[0].row, end_column=6)
            sheet.merge_cells(start_row=row[0].row, start_column=7, end_row=row[0].row, end_column=10)
            sheet.merge_cells(start_row=row[0].row, start_column=11, end_row=row[0].row, end_column=14)
            sheet.merge_cells(start_row=row[0].row, start_column=15, end_row=row[0].row, end_column=16)
            for cell in row:
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        elif check_drug:
            sheet.row_dimensions[row[0].row].height = 50
            sheet.merge_cells(start_row=row[0].row, start_column=2, end_row=row[0].row, end_column=11)
            sheet.merge_cells(start_row=row[0].row, start_column=12, end_row=row[0].row, end_column=16)
            for cell in row:
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        elif contains_postal_code:
            sheet.row_dimensions[row[0].row].height = 80
            sheet.merge_cells(start_row=row[0].row, start_column=1, end_row=row[0].row, end_column=4)
            for cell in row:
                cell.alignment = Alignment(vertical='bottom', horizontal='left', wrap_text=True)
            sheet.merge_cells(start_row=row[0].row, start_column=10, end_row=row[0].row, end_column=16)
            for cell in row:
                cell.alignment = Alignment(vertical='bottom', wrap_text=True)
        elif contains_tohoku:
            sheet.row_dimensions[row[0].row].height = 125
            sheet.merge_cells(start_row=row[0].row, start_column=1, end_row=row[0].row, end_column=16)
            for cell in row:
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        elif contains_expert_panel:
            sheet.row_dimensions[row[0].row].height = 100
        elif any('署名' in str(cell.value) for cell in row):
            sheet.row_dimensions[row[0].row].height = 50
            sheet.merge_cells(start_row=row[0].row, start_column=1, end_row=row[0].row, end_column=6)
            for cell in row:
                cell.alignment = Alignment(vertical='bottom', horizontal='left', wrap_text=True)
        elif any('年　　　月　　　日' in str(cell.value) for cell in row):
            sheet.row_dimensions[row[0].row].height = 50
            sheet.merge_cells(start_row=row[0].row, start_column=1, end_row=row[0].row, end_column=4)
            for cell in row:
                cell.alignment = Alignment(vertical='bottom', horizontal='right', wrap_text=True)

    for row in sheet.iter_rows():
        if 'evidence' not in str(row[16].value):
            count = sum(cell.value.count('::') for cell in row if cell.value and isinstance(cell.value, str) and '::' in cell.value)
            if count > 1:
                sheet.row_dimensions[row[0].row].height = count * 25
        for cell in row:
            if cell.value and isinstance(cell.value, str) and ('FDA' in cell.value or 'PMDA' in cell.value):
                fda_pmda_count = cell.value.count('FDA') + cell.value.count('PMDA')
                sheet.row_dimensions[row[0].row].height = fda_pmda_count * 25
                break

    # 16行目以下の17列目を削除
    for row in range(16, sheet.max_row + 1):
        sheet.cell(row=row, column=17).value = None
    
    # セルを結合
    sheet.merge_cells(start_row=3, start_column=17, end_row=15, end_column=26)
    sheet.merge_cells(start_row=3, start_column=27, end_row=7, end_column=32)
    sheet.merge_cells(start_row=9, start_column=27, end_row=15, end_column=32)

    # 'バリアントの評価'を含むセルから改ページ
    """
    for row in sheet.iter_rows(min_row=16, max_col=32):
        if any('バリアントの評価' in str(cell.value) for cell in row):
            sheet.row_dimensions[row[0].row].page_break = openpyxl.worksheet.dimensions.Break.ROW
    """
    # 印刷範囲は'contains_postal_code'を含むセルまで
    sheet.print_area = f'A1:AF{sheet.max_row}'
    
    output_stream = BytesIO()
    wb.save(output_stream)
    wb.close()
    output_stream.seek(0)

    return output_stream


def excel_fasttrack(analysis_type, pdf_data, template_path, date, ep_institution, ep_department, ep_responsible, ep_contact, ep_tel):
    current_dir = os.getcwd()
    try:
        # pdf_data が bytes の場合は BytesIO でラップ
        pdf_file_like = io.BytesIO(pdf_data)
        
        with pdfplumber.open(pdf_file_like) as pdf:
            all_text = "\n".join(
                page.extract_text() for page in pdf.pages if page.extract_text()
            )
    except Exception as e:
        st.error(f"PDFの読み込み中にエラーが発生しました: {e}")
        st.stop()

    lines = all_text.strip().split("\n")

    marker_ranges = ReadRange.FASTTRACK

    combined_json = {}

    for section in marker_ranges:
        start_idx, end_idx = -1, -1
        for i, line in enumerate(lines):
            if section["start_marker"] in line and start_idx == -1:
                start_idx = i + 1
            elif section["end_marker"] in line and start_idx != -1:
                end_idx = i
                break

        if start_idx != -1 and end_idx != -1:
            extracted_lines = lines[start_idx:end_idx]
            parsed_data = []

            for line in extracted_lines:
                items = re.split(r"\s+", line.strip())
                items = [item for item in items if item]

                record = {}
                for i, item in enumerate(items):
                    if i < len(section["columns"]):
                        record[section["columns"][i]] = item
                    else:
                        record[f"Extra_{i+1}"] = item

                parsed_data.append(record)

            combined_json[section["name"]] = parsed_data
        else:
            st.warning(f"{section['name']} の範囲が見つかりませんでした。")

    if combined_json:
        json_str = json.dumps(combined_json, ensure_ascii=False, indent=2)
        
        # Excel output processing
        df_disease = pd.DataFrame(combined_json.get("Disease", []))
        df_disease = df_disease.apply(lambda x: ' '.join(x.dropna().astype(str)), axis=1)
        disease = ' '.join(df_disease.tolist())
        try:
            disease_1 = disease.split('疾患名 ')[1].split(' エビデンスレベル')[0]
            disease_2 = disease.split('エビデンスレベル付与対象の ')[1]
        except IndexError:
            disease_1 = ""
            disease_2 = ""

        df_snv_indel = pd.DataFrame(combined_json.get("SNV_Indel", []))
        try:
            df_snv_indel = df_snv_indel[Columns.HEMESIGHT_FASTTRACK_SNV]
        except KeyError:
            df_snv_indel = pd.DataFrame(columns=Columns.HEMESIGHT_FASTTRACK_SNV)
        df_sv = pd.DataFrame(combined_json.get("SV", []))
        try:
            df_sv = df_sv[Columns.HEMESIGHT_FASTTRACK_CNV]
        except KeyError:
            df_sv = pd.DataFrame(columns=Columns.HEMESIGHT_FASTTRACK_CNV)

        wb = openpyxl.load_workbook(template_path)
        sheet = wb["FTReport"]

        # Insert expert panel inputs
        add_logo(current_dir, sheet, cell='A1')
        sheet['D2'] = analysis_type
        sheet['A24'] = '報告書作成日：' + date.strftime('%Y年%m月%d日')
        sheet['J2'] = ep_institution
        sheet['J4'] = ep_department + ' / ' + ep_responsible
        sheet['J25'] = ep_department + ' ' + ep_responsible
        sheet['H28'] = ep_contact + '\n' + '電話番号 ' + ep_tel + '\n' + ep_institution + ' ' + ep_department
        sheet['J7'] = disease_1
        sheet['J8'] = disease_2

        start_row = 14
        insert_row(wb, df_snv_indel, sheet_name='FTReport', start_row=start_row, start_col='A', end_col='P')
        for col_pos in [1, 3, 4, 6, 8, 9, 10, 12, 13, 14]:
            if col_pos < len(df_snv_indel.columns):
                df_snv_indel.insert(col_pos, f"Col_{col_pos}", "")

        # Insert SV data
        star_row_sv = 18 + len(df_snv_indel) - 1
        insert_row(wb, df_sv, sheet_name='FTReport', start_row=star_row_sv, start_col='A', end_col='P')
        for col_pos in [1, 3, 4, 6, 8, 9, 11, 12, 14]:
            if col_pos < len(df_sv.columns):
                df_sv.insert(col_pos, f"Col_{col_pos}", "")

        for df_section, start_row in [(df_snv_indel, start_row), (df_sv, star_row_sv)]:
            for r_idx, row in enumerate(dataframe_to_rows(df_section, index=False, header=False), start_row):
                for c_idx, value in enumerate(row, 1):
                    sheet.cell(row=r_idx, column=c_idx, value=value)
                    sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(wrap_text=False, vertical='top')

        # 空行削除
        # delete_blank_lines(sheet)

        # Adjust formatting
        for row in sheet.iter_rows():
            row_values = [str(cell.value) if cell.value else "" for cell in row]
            contains_expert_panel = any('Fast-track持ち回り協議結果報告書' in val for val in row_values)
            contains_summary = any('以上の遺伝子異常を確認しました。' in val for val in row_values)
            contains_postal_code = any('〒' in val for val in row_values)

            if contains_postal_code:
                sheet.row_dimensions[row[0].row].height = 100
                sheet.merge_cells(start_row=row[0].row, start_column=8, end_row=row[0].row, end_column=16)
                for cell in row:
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
            elif contains_summary:
                sheet.row_dimensions[row[0].row].height = 100
                sheet.merge_cells(start_row=row[0].row, start_column=1, end_row=row[0].row, end_column=16)
                for cell in row:
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
            elif contains_expert_panel:
                sheet.row_dimensions[row[0].row].height = 100

        # 印刷範囲
        sheet.print_area = f'A1:P{sheet.max_row}'

        output_stream = BytesIO()
        wb.save(output_stream)
        wb.close()
        output_stream.seek(0)

    return output_stream
    
    

def excel_foundationone(analysis_type, output_stream, date, ep_institution, ep_department, ep_responsible, ep_contact, ep_tel):
    current_dir = os.getcwd()

    xlsx = pd.read_excel(output_stream, sheet_name=None)
    df_short = xlsx['SNV_Indel']
    df_cnv = xlsx['CNV']
    df_fusion = xlsx['Fusion']
    df_msi = xlsx['MSI']
    
    if not df_short.empty:
        df_short = df_short.copy()
        df_short['alternateAlleleFrequency'] = df_short.apply(lambda x: x['alternateAlleleReadDepth'] / x['totalReadDepth'] if pd.notna(x['alternateAlleleReadDepth']) and pd.notna(x['totalReadDepth']) and x['totalReadDepth'] != 0 else None, axis=1)
        df_short['GeneBe_ClinVar_Germline'] = df_short['GeneBe_ClinVar_Germline'].fillna('').astype(str) + ' ' + df_short['GeneBe_ClinVar_Germline_Status'].fillna('').astype(str)
        df_short['GeneBe_ClinVar_Somatic'] = df_short['GeneBe_ClinVar_Somatic'].fillna('').astype(str) + ' ' + df_short['GeneBe_ClinVar_Somatic_Status'].fillna('').astype(str)   
        df_short.loc[df_short['GeneBe_ClinVar_Germline'].isin(['§']), 'GeneBe_ClinVar_Germline'] = '(' + df_short['GeneBe_ClinVar_Submission_Summary'].astype(str) + ')'
        df_snv = df_short[Columns.SNV_INDEL].copy()
        df_snv['alternateAlleleFrequency'] = df_snv['alternateAlleleFrequency'].apply(lambda x: f"{x * 100:.1f}%" if pd.notna(x) else '')
        df_snv = df_snv.sort_values(by=['status', 'geneSymbol'])
    
    if not df_cnv.empty:
        df_cnv = df_cnv[Columns.FOUNDATION_CNV].copy()
        df_cnv = df_cnv.sort_values(by=['status', 'geneSymbol'])
        
    if not df_fusion.empty:
        df_fusion = df_fusion[Columns.FOUNDATION_FUSION].copy()
        df_fusion = df_fusion.sort_values(by=['status', 'description'])
    
    df_germline = df_short[Columns.FOUNDATION_GERMLINE].copy()
    base_path = os.path.dirname(__file__)
    df_pgpv = pd.read_csv(os.path.join(base_path, Database.PGPV_PATH))
    df_germline = pd.merge(df_germline, df_pgpv, on='geneSymbol', how='left')
    if analysis_type == 'FoundationOne':
        if df_germline['cdsChange'].str.contains('>', na=False).any():
            df_germline = df_germline[df_germline['alternateAlleleFrequency'] > df_germline['SNV']]
        elif df_germline['cdsChange'].str.contains('del|ins', na=False).any():
            df_germline = df_germline[df_germline['alternateAlleleFrequency'] > df_germline['Indel']]
        if (df_msi['status'] != 'MSI').any():
            df_germline = df_germline[~df_germline['geneSymbol'].isin(['MLH1', 'MSH2', 'MSH6', 'PMS2'])]
        else:
            pass
    elif analysis_type == 'FoundationOne Liquid':
        df_germline = df_germline[df_germline['alternateAlleleFrequency'].astype(float) > 0.3]
    df_germline = df_germline[['geneSymbol', 'aminoAcidsChange', 'Comment']].copy()
    df_germline['Merged'] = df_germline['geneSymbol'] + ' ' + df_germline['aminoAcidsChange'] + ' ' + df_germline['Comment']
    df_germline = df_germline[['Merged']]
    
    
    ####################################
    wb = openpyxl.load_workbook(output_stream)
    sheet = wb['Summary']
    
    # 基本情報の入力
    add_logo(current_dir, sheet, cell='A1')
    sheet['D2'] = analysis_type
    sheet['D4'] = date.strftime('%Y年%m月%d日')
    sheet['J2'] = ep_institution
    sheet['J4'] = f'{ep_department} / {ep_responsible}'
    sheet['A54'] = '報告書作成日：' + date.strftime('%Y年%m月%d日')
    sheet['A55'] = f'{ep_institution} {ep_department}'
    sheet['J55'] = ep_responsible
    sheet['J58'] = f'{ep_institution} {ep_department}\n{ep_contact}\n電話番号 {ep_tel}'

    start_row_snv = 22
    insert_row(wb, df_snv, sheet_name='Summary', start_row=start_row_snv, start_col='A', end_col='O')
    for insert_pos in [2, 3, 7]:
        df_snv.insert(insert_pos, chr(96 + insert_pos), '')
    start_row_cnv = start_row_snv + 5 + len(df_snv) - 1 if len(df_snv) > 1 else start_row_snv + 5
    insert_row(wb, df_cnv, sheet_name='Summary', start_row=start_row_cnv, start_col='A', end_col='O')
    for insert_pos in [2, 4, 7, 8, 10]:
        col_name = chr(96 + insert_pos)
        if 0 <= insert_pos <= df_cnv.shape[1]:
            df_cnv.insert(insert_pos, col_name, '')
        else:
            df_cnv[col_name] = ''
    start_row_fusion = start_row_cnv + 4 + len(df_cnv) - 1 if len(df_cnv) > 1 else start_row_cnv + 4
    insert_row(wb, df_fusion, sheet_name='Summary', start_row=start_row_fusion, start_col='A', end_col='O')
    for insert_pos in [1, 2, 3, 4, 7, 8, 10]:
        col_name = chr(96 + insert_pos)      
        if 0 <= insert_pos <= df_fusion.shape[1]:
            df_fusion.insert(insert_pos, col_name, '')
        else:
            df_fusion[col_name] = ''
    start_row_germline = start_row_fusion+ 15 + len(df_fusion) - 1 if len(df_fusion) > 1 else start_row_fusion + 15
    insert_row(wb, df_germline, sheet_name='Summary', start_row=start_row_germline, start_col='A', end_col='O')
    for df_section, start_row in [(df_snv, start_row_snv), (df_cnv, start_row_cnv), (df_fusion, start_row_fusion), (df_germline, start_row_germline)]:
        for r_idx, row in enumerate(dataframe_to_rows(df_section, index=False, header=False), start_row):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)
                sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(wrap_text=False, vertical='top')

    # 条件に応じた行処理
    for row in sheet.iter_rows():
        values = [cell.value for cell in row]
        r = row[0].row

        if any('〒' in str(v) for v in values):
            sheet.row_dimensions[r].height = 100
            sheet.merge_cells(start_row=r, start_column=10, end_row=r, end_column=15)
        elif 'Tohoku' in str(values):
            sheet.row_dimensions[r].height = 150
            sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=15)
        elif 'エキスパートパネルレポート' in str(values):
            sheet.row_dimensions[r].height = 100
        elif '生殖細胞系列由来' in str(values) :
            sheet.row_dimensions[r].height = 100
            sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=15)
            sheet.cell(row=r, column=1).alignment = Alignment(vertical='center', wrap_text=True)

    # 固定セル結合
    sheet.merge_cells(start_row=3, start_column=17, end_row=15, end_column=26)
    sheet.merge_cells(start_row=3, start_column=27, end_row=7, end_column=32)
    sheet.merge_cells(start_row=9, start_column=27, end_row=15, end_column=32)

    # 印刷範囲設定
    sheet.print_area = f'A1:AF{sheet.max_row}'
    # output_stream.seek(0)
    
    ####################################
    sheet = wb['For_pts'] 

    # 基本情報の入力
    add_logo(current_dir, sheet, cell='A1')
    sheet['C2'] = analysis_type
    sheet['C4'] = ep_institution
    sheet['F5'] = f'{ep_department} / {ep_responsible}'
    sheet['A37'] = '報告書作成日：' + date.strftime('%Y年%m月%d日')
    sheet['A38'] = f'{ep_institution} {ep_department}'
    sheet['F38'] = ep_responsible
    sheet['E41'] = f'{ep_institution} {ep_department}\n{ep_contact}\n電話番号 {ep_tel}'
    
    # 必要な列のみ抽出
    df_patient = df_short[['geneSymbol', 'aminoAcidsChange']].copy()

    # 空の列を挿入（必要に応じて）
    start_row_patient = 26
    insert_row(wb, df_patient, sheet_name='For_pts', start_row=start_row_patient, start_col='A', end_col='F')
    for insert_pos in [1]:
        df_patient.insert(insert_pos, f'column_{chr(96 + insert_pos)}', '')
    
    # df_germline
    start_row_germline = 30 + len(df_patient) - 1
    insert_row(wb, df_germline, sheet_name='For_pts', start_row=start_row_germline, start_col='A', end_col='F')

    for df_section, start_row in [(df_patient, start_row_patient), (df_germline, start_row_germline)]:
        for r_idx, row in enumerate(dataframe_to_rows(df_section, index=False, header=False), start_row):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)
                sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(wrap_text=False, vertical='top')

    # 空行削除
    # delete_blank_lines(sheet)

    # 条件に応じた行処理
    for row in sheet.iter_rows():
        values = [cell.value for cell in row]
        r = row[0].row
        if any('〒' in str(v) for v in values):
            sheet.row_dimensions[r].height = 100
            sheet.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        elif 'がん遺伝子パネル検査説明書' in str(values):
            sheet.row_dimensions[r].height = 100
        elif '生殖細胞系列由来' in str(values) :
            sheet.row_dimensions[r].height = 100
            sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            sheet.cell(row=r, column=1).alignment = Alignment(vertical='center', wrap_text=True)        

    # 印刷範囲を設定（A1からF列までの最大行）
    sheet.print_area = f'A1:F{sheet.max_row}'

    # 新しい出力ストリームを作成し保存
    output_stream = BytesIO()
    wb.save(output_stream)
    wb.close()
    output_stream.seek(0)
    return output_stream


def excel_genminetop(analysis_type, output_stream, date, ep_institution, ep_department, ep_responsible, ep_contact, ep_tel):
    current_dir = os.getcwd()

    xlsx = pd.read_excel(output_stream, sheet_name=None)
    df_short = xlsx['SNV_Indel']
    df_cnv = xlsx['CNV']
    df_fusion = xlsx['Fusion']
    df_germline = xlsx['Germline']

    ####################################    
    wb = openpyxl.load_workbook(output_stream)
    sheet = wb['Summary']
                    
    # 基本情報の入力
    add_logo(current_dir, sheet, cell='A1')
    sheet['D2'] = analysis_type
    sheet['D4'] = date.strftime('%Y年%m月%d日')
    sheet['J2'] = ep_institution
    sheet['J4'] = f'{ep_department} / {ep_responsible}'
    sheet['A55'] = '報告書作成日：' + date.strftime('%Y年%m月%d日')
    sheet['A56'] = f'{ep_institution} {ep_department}'
    sheet['J56'] = ep_responsible
    sheet['H59'] = f'{ep_institution} {ep_department}\n{ep_contact}\n電話番号 {ep_tel}'
    
    def process_variant_df(df):
        if df.empty:
            return
        df = df.copy()
        # 欠損補完（aminoAcidsChange）
        df['aminoAcidsChange'] = df['aminoAcidsChange'].fillna(df['cdsChange'])

        # 分数形式からパーセントへ変換
        df['alternateAlleleReadDepth'] = df['alternateAlleleFrequency'].str.split('/').str[0].astype(float)
        df['totalReadDepth'] = df['alternateAlleleFrequency'].str.split('/').str[1].astype(float)
        df['alternateAlleleFrequency'] = df.apply(
            lambda x: f"{x['alternateAlleleReadDepth'] / x['totalReadDepth'] * 100:.1f}%" 
            if pd.notna(x['alternateAlleleReadDepth']) and pd.notna(x['totalReadDepth']) else '-', axis=1)

        # ClinVar 情報を統合
        df['GeneBe_ClinVar_Germline'] = df['GeneBe_ClinVar_Germline'].fillna('').astype(str) + ' ' + df['GeneBe_ClinVar_Germline_Status'].fillna('').astype(str)
        df['GeneBe_ClinVar_Somatic'] = df['GeneBe_ClinVar_Somatic'].fillna('').astype(str) + ' ' + df['GeneBe_ClinVar_Somatic_Status'].fillna('').astype(str)

        # '§' の置換
        mask = df['GeneBe_ClinVar_Germline'].str.strip() == '§'
        df.loc[mask, 'GeneBe_ClinVar_Germline'] = '(' + df.loc[mask, 'GeneBe_ClinVar_Submission_Summary'].astype(str) + ')'

        # 出力用DataFrameの作成
        df_out = df[Columns.SNV_INDEL].copy()

        # ソート
        df_out = df_out.sort_values(by=['status', 'geneSymbol'])

        # 空列挿入
        for insert_pos in [2, 3, 7]:
            df_out.insert(insert_pos, chr(96 + insert_pos), '')
        
        return df_out
                
    df_germline = process_variant_df(df_germline)
    df_short = process_variant_df(df_short)

    if not df_cnv.empty:
        df_cnv = df_cnv[Columns.GENMINE_CNV].copy()
        for insert_pos in [2, 4, 5, 7, 8, 9, 10]:
            col_name = chr(96 + insert_pos)
            if 0 <= insert_pos <= df_cnv.shape[1]:
                df_cnv.insert(insert_pos, col_name, '')
            else:
                df_cnv[col_name] = ''

    if not df_fusion.empty:
        df_fusion['Fusion_Gene'] = df_fusion['gene_1'] + ' - ' + df_fusion['gene_2']
        df_fusion['Breakpoint_1'] = df_fusion['chr_1'] + ":" + df_fusion['pos_1'].astype(str)
        df_fusion['Breakpoint_2'] = df_fusion['chr_2'] + ":" + df_fusion['pos_2'].astype(str)
        df_fusion['Breakpoint'] = df_fusion['Breakpoint_1'] + ' - ' + df_fusion['Breakpoint_2']
        
        df_fusion = df_fusion[Columns.GENMINE_FUSION].copy()
        for insert_pos in [1, 3, 4, 5, 7, 8, 10]:
            col_name = chr(96 + insert_pos)
            if 0 <= insert_pos <= df_fusion.shape[1]:
                df_fusion.insert(insert_pos, col_name, '')
            else:
                df_fusion[col_name] = ''
    
    if not df_germline.empty:            
        base_path = os.path.dirname(__file__)
        df_pgpv = pd.read_csv(os.path.join(base_path, Database.PGPV_PATH))
        df_germline_comment = pd.merge(df_germline, df_pgpv, on='geneSymbol', how='left')
        df_germline_comment = df_germline_comment[['geneSymbol', 'aminoAcidsChange', 'Comment_GPV']].copy()
        df_germline_comment['Merged'] = df_germline_comment['geneSymbol'] + ' ' + df_germline_comment['aminoAcidsChange'] + ' ' + df_germline_comment['Comment_GPV']
        df_germline_comment = df_germline_comment[['Merged']]

    start_row_germline = 17
    insert_row(wb, df_germline, sheet_name='Summary', start_row=start_row_germline, start_col='A', end_col='O')
    start_row_short = start_row_germline + 6 + len(df_germline) - 1 if len(df_germline) > 1 else start_row_germline + 6
    insert_row(wb, df_short, sheet_name='Summary', start_row=start_row_short, start_col='A', end_col='O')
    
    df_short_len = len(df_short) if df_short is not None else 0
    start_row_cnv = start_row_short + 5 + df_short_len - 1 if df_short_len > 1 else start_row_short + 5
    insert_row(wb, df_cnv, sheet_name='Summary', start_row=start_row_cnv, start_col='A', end_col='O')
    start_row_fusion = start_row_cnv + 4 + len(df_cnv) - 1 if len(df_cnv) > 1 else start_row_cnv + 4
    insert_row(wb, df_fusion, sheet_name='Summary', start_row=start_row_fusion, start_col='A', end_col='O')
    start_row_germline_comment = start_row_fusion + 15 + len(df_fusion) - 1 if len(df_fusion) > 1 else start_row_fusion + 15
    insert_row(wb, df_germline_comment, sheet_name='Summary', start_row=start_row_germline_comment, start_col='A', end_col='O')
    
    for df_section, start_row in [(df_germline, start_row_germline), (df_short, start_row_short), (df_cnv, start_row_cnv), 
                                  (df_fusion, start_row_fusion), (df_germline_comment, start_row_germline_comment)]:
        if df_section is not None and not df_section.empty:
            for r_idx, row in enumerate(dataframe_to_rows(df_section, index=False, header=False), start_row):
                for c_idx, value in enumerate(row, 1):
                    sheet.cell(row=r_idx, column=c_idx, value=value)
                    sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(wrap_text=False, vertical='top')

    # 空行削除
    # delete_blank_lines(sheet)

    # 条件に応じた行処理
    for row in sheet.iter_rows():
        values = [cell.value for cell in row]
        r = row[0].row
        if any('〒' in str(v) for v in values):
            sheet.row_dimensions[r].height = 100
            sheet.merge_cells(start_row=r, start_column=8, end_row=r, end_column=15)
        elif 'Tohoku' in str(values):
            sheet.row_dimensions[r].height = 150
            sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=15)
        elif 'エキスパートパネルレポート' in str(values):
            sheet.row_dimensions[r].height = 100
        elif '生殖細胞系列由来' in str(values) :
            sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=15)
            sheet.cell(row=r, column=1).alignment = Alignment(vertical='center', wrap_text=True)
        elif 'ACMG' in str(values):
            sheet.row_dimensions[r].height = 50       

    # 固定セル結合
    sheet.merge_cells(start_row=3, start_column=17, end_row=15, end_column=26)
    sheet.merge_cells(start_row=3, start_column=27, end_row=7, end_column=32)
    sheet.merge_cells(start_row=9, start_column=27, end_row=15, end_column=32)
                        
    # 印刷範囲を設定（A1からF列までの最大行）
    sheet.print_area = f'A1:AF{sheet.max_row}'

    ####################################
    sheet = wb['For_pts'] 

    # 基本情報の入力
    add_logo(current_dir, sheet, cell='A1')
    sheet['C2'] = analysis_type
    sheet['C4'] = ep_institution
    sheet['F5'] = f'{ep_department} / {ep_responsible}'
    sheet['A38'] = '報告書作成日：' + date.strftime('%Y年%m月%d日')
    sheet['A39'] = f'{ep_institution} {ep_department}'
    sheet['F39'] = ep_responsible
    sheet['E42'] = f'{ep_institution} {ep_department}\n{ep_contact}\n電話番号 {ep_tel}'

    # 必要な列のみ抽出
    if df_short is not None and not df_short.empty:
        df_patient = df_short[['geneSymbol', 'aminoAcidsChange']].copy()
    else:
        df_patient = pd.DataFrame(columns=['geneSymbol', 'aminoAcidsChange'])

    # df_patient
    start_row_patient = 26
    insert_row(wb, df_patient, sheet_name='For_pts', start_row=start_row_patient, start_col='A', end_col='F')
    start_row_germline = start_row_patient + 5 + len(df_patient) - 1 if len(df_patient) > 1 else start_row_patient + 5
    insert_row(wb, df_germline, sheet_name='For_pts', start_row=start_row_germline, start_col='A', end_col='F')
    for insert_pos in [1]:
        df_patient.insert(insert_pos, f'column_{chr(96 + insert_pos)}', '')

    for df_section, start_row in [(df_patient, start_row_patient), (df_germline_comment, start_row_germline)]:
        for r_idx, row in enumerate(dataframe_to_rows(df_section, index=False, header=False), start_row):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)
                sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(wrap_text=False, vertical='top')

    # 空行削除
    # delete_blank_lines(sheet)

    # 条件に応じた行処理
    for row in sheet.iter_rows():
        values = [cell.value for cell in row]
        r = row[0].row
        if any('〒' in str(v) for v in values):
            sheet.row_dimensions[r].height = 100
            sheet.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        elif 'がん遺伝子パネル検査説明書' in str(values):
            sheet.row_dimensions[r].height = 100
        elif '生殖細胞系列由来' in str(values) :
            sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            sheet.cell(row=r, column=1).alignment = Alignment(vertical='center', wrap_text=True)        

    # 印刷範囲を設定（A1からF列までの最大行）
    sheet.print_area = f'A1:F{sheet.max_row}'

    # 新しい出力ストリームを作成し保存
    output_stream = BytesIO()
    wb.save(output_stream)
    wb.close()
    output_stream.seek(0)
    return output_stream


def excel_guardant360(analysis_type, output_stream, date, ep_institution, ep_department, ep_responsible, ep_contact, ep_tel):
    current_dir = os.getcwd()
    
    xlsx = pd.read_excel(output_stream, sheet_name=None)
    df_snv = xlsx['SNV']
    df_indel = xlsx['Indels']
    
    if not df_snv.empty and not df_indel.empty:
        df_short = pd.concat([df_snv, df_indel], ignore_index=True)
    elif not df_snv.empty and df_indel.empty:
        df_short = df_snv.copy()
    elif df_snv.empty and not df_indel.empty:
        df_short = df_indel.copy()
    
    df_cnv = xlsx['CNAs']
    df_fusion = xlsx['Fusions']
    df_msi = xlsx['MSI']

    if not df_short.empty:
        df_germline = df_short.copy()
        base_path = os.path.dirname(__file__)
        df_pgpv = pd.read_csv(os.path.join(base_path, Database.PGPV_PATH))
        df_germline = pd.merge(df_germline, df_pgpv, on='geneSymbol', how='left')
        df_germline = df_germline[df_germline['alternateAlleleFrequency'].astype(float) > 30]

        if (df_msi['msi_status'] != 'MSI-H').any():
            df_germline = df_germline[~df_germline['geneSymbol'].isin(['MLH1', 'MSH2', 'MSH6', 'PMS2'])]
        else:
            pass
        # 欠損補完（aminoAcidsChange）
        df_germline['aminoAcidsChange'] = df_germline['aminoAcidsChange'].fillna(df_germline['cdsChange'])
        
        df_germline = df_germline[['geneSymbol', 'aminoAcidsChange', 'Comment']].copy()
        df_germline['Merged'] = df_germline['geneSymbol'] + ' ' + df_germline['aminoAcidsChange'] + ' ' + df_germline['Comment']
        df_germline = df_germline[['Merged']]
    
    df_short = df_short.copy()
    df_short['aminoAcidsChange'] = df_short['aminoAcidsChange'].fillna(df_short['cdsChange'])
    df_short['GeneBe_ClinVar_Germline'] = df_short['GeneBe_ClinVar_Germline'].fillna('').astype(str) + ' ' + df_short['GeneBe_ClinVar_Germline_Status'].fillna('').astype(str)
    df_short['GeneBe_ClinVar_Somatic'] = df_short['GeneBe_ClinVar_Somatic'].fillna('').astype(str) + ' ' + df_short['GeneBe_ClinVar_Somatic_Status'].fillna('').astype(str)
    mask = df_short['GeneBe_ClinVar_Germline'].str.strip() == '§'
    df_short.loc[mask, 'GeneBe_ClinVar_Germline'] = '(' + df_short.loc[mask, 'GeneBe_ClinVar_Submission_Summary'].astype(str) + ')'
    df_short = df_short[Columns.SNV_INDEL].copy()
    df_short = df_short.sort_values(by=['status', 'geneSymbol'])
    
    # df_cnv
    df_cnv = df_cnv.copy()
    try:
        df_cnv = df_cnv[Columns.GUARDANT_CNV].copy()
    except KeyError:
        df_cnv = pd.DataFrame(columns=Columns.GUARDANT_CNV)

    ####################################
    wb = openpyxl.load_workbook(output_stream)
    sheet = wb['Summary']

    # 基本情報の入力
    add_logo(current_dir, sheet, cell='A1')
    sheet['D2'] = analysis_type
    sheet['D4'] = date.strftime('%Y年%m月%d日')
    sheet['J2'] = ep_institution
    sheet['J4'] = f'{ep_department} / {ep_responsible}'
    sheet['A53'] = '報告書作成日：' + date.strftime('%Y年%m月%d日')
    sheet['A54'] = f'{ep_institution} {ep_department}'
    sheet['J54'] = ep_responsible
    sheet['H57'] = f'{ep_institution} {ep_department}\n{ep_contact}\n電話番号 {ep_tel}'

    # df_short
    start_row_short = 21
    insert_row(wb, df_short, sheet_name='Summary', start_row=start_row_short, start_col='A', end_col='O')
    for insert_pos in [2, 3, 7]:
        df_short.insert(insert_pos, chr(96 + insert_pos), '')

    # df_cnv
    start_row_cnv = 26 + len(df_short) - 1
    insert_row(wb, df_cnv, sheet_name='Summary', start_row=start_row_cnv, start_col='A', end_col='O')
    for insert_pos in [2, 4, 5, 7, 8, 10]:
        df_cnv.insert(insert_pos, chr(96 + insert_pos), '')
        
    # df_germline
    start_row_germline = 45 + len(df_short) + len(df_cnv) - 2
    insert_row(wb, df_germline, sheet_name='Summary', start_row=start_row_germline, start_col='A', end_col='O')
    
    for df_section, start_row in [(df_short, start_row_short), (df_cnv, start_row_cnv), (df_germline, start_row_germline)]:
        for r_idx, row in enumerate(dataframe_to_rows(df_section, index=False, header=False), start_row):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)
                sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(wrap_text=False, vertical='top')

    # 空行削除
    # delete_blank_lines(sheet)

    # 条件に応じた行処理
    for row in sheet.iter_rows():
        values = [cell.value for cell in row]
        r = row[0].row
        if any('〒' in str(v) for v in values):
            sheet.row_dimensions[r].height = 100
            sheet.merge_cells(start_row=r, start_column=8, end_row=r, end_column=15)
            sheet.cell(row=r, column=8).alignment = Alignment(vertical='center', wrap_text=True)
        elif 'Tohoku' in str(values):
            sheet.row_dimensions[r].height = 180
            sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=15)
        elif 'エキスパートパネルレポート' in str(values):
            sheet.row_dimensions[r].height = 100
        elif '生殖細胞系列由来' in str(values) :
            sheet.row_dimensions[r].height = 80
            sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=15)
            sheet.cell(row=r, column=1).alignment = Alignment(vertical='center', wrap_text=True)
        elif 'ACMG' in str(values):
            sheet.row_dimensions[r].height = 50       

    # 固定セル結合
    sheet.merge_cells(start_row=3, start_column=17, end_row=15, end_column=26)
    sheet.merge_cells(start_row=3, start_column=27, end_row=7, end_column=32)
    sheet.merge_cells(start_row=9, start_column=27, end_row=15, end_column=32)
                        
    # 印刷範囲を設定（A1からF列までの最大行）
    sheet.print_area = f'A1:AF{sheet.max_row}'    

    ####################################
    sheet = wb['For_pts']  # 'For_pts'シートを選択

    # 必要な列のみ抽出
    df_patient = df_short[['geneSymbol', 'aminoAcidsChange']].copy()

    # 基本情報の入力
    add_logo(current_dir, sheet, cell='A1')
    sheet['C2'] = analysis_type
    sheet['C4'] = ep_institution
    sheet['F5'] = f'{ep_department} / {ep_responsible}'
    sheet['A33'] = '報告書作成日：' + date.strftime('%Y年%m月%d日')
    sheet['A34'] = f'{ep_institution} {ep_department}'
    sheet['F34'] = ep_responsible
    sheet['E37'] = f'{ep_institution} {ep_department}\n{ep_contact}\n電話番号 {ep_tel}'

    # 空の列を挿入（必要に応じて）
    start_row_patient = 22
    insert_row(wb, df_patient, sheet_name='For_pts', start_row=start_row_patient, start_col='A', end_col='F')
    for insert_pos in [1]:
        df_patient.insert(insert_pos, f'column_{chr(96 + insert_pos)}', '')
        
    # df_germline
    start_row_germline = 26 + len(df_patient) - 1
    insert_row(wb, df_germline, sheet_name='For_pts', start_row=start_row_germline, start_col='A', end_col='F')

    for df_section, start_row in [(df_patient, start_row_patient), (df_germline, start_row_germline)]:
        for r_idx, row in enumerate(dataframe_to_rows(df_section, index=False, header=False), start_row):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)
                sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(wrap_text=False, vertical='top')

    # 空行削除
    # delete_blank_lines(sheet)

    # 条件に応じた行処理
    for row in sheet.iter_rows():
        values = [cell.value for cell in row]
        r = row[0].row
        if any('〒' in str(v) for v in values):
            sheet.row_dimensions[r].height = 100
            sheet.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        elif 'がん遺伝子パネル検査説明書' in str(values):
            sheet.row_dimensions[r].height = 100
        elif '生殖細胞系列由来' in str(values) :
            sheet.row_dimensions[r].height = 80
            sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            sheet.cell(row=r, column=1).alignment = Alignment(vertical='center', wrap_text=True)        
                    
    # 印刷範囲を設定（A1からF列までの最大行）
    sheet.print_area = f'A1:F{sheet.max_row}'

    output_stream = BytesIO()
    wb.save(output_stream)
    wb.close()
    output_stream.seek(0)    
    
    return output_stream